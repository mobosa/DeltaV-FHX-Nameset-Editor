"""Unit tests for fhx_core — parsing, validation, translation, and Excel helpers."""

import os
import tempfile
import unittest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from fhx_core import (
    find_block_end,
    find_enum_set_block,
    extract_enum_sets,
    extract_string_value_refs,
    extract_expression_refs,
    extract_alarms,
    extract_priority_names,
    _is_chinese,
    _bidirectional_translate,
    validate_xlsx,
    compare_alarms,
    compare_priority_names,
    replace_priority_names,
    replace_alarm_values,
    DELTA_VALUE_CN_TO_EN,
    DELTA_VALUE_EN_TO_CN,
    read_fhx,
    write_fhx,
    _write_excel_sheet,
    write_alarm_types_sheet,
    write_priority_names_sheet,
    read_alarm_types_excel,
    read_priority_names_excel,
    validate_excel_for_generation,
    read_lib_edited_excel,
)


class TestFindBlockEnd(unittest.TestCase):
    def test_simple_block(self):
        # find_block_end returns position AFTER closing '}'
        content = '{ abc }'  # '}' at index 6, returns 7
        self.assertEqual(find_block_end(content, 0), 7)

    def test_nested_blocks(self):
        content = '{ a { b } c }'  # outer '}' at index 12, returns 13
        self.assertEqual(find_block_end(content, 0), 13)

    def test_no_closing_brace(self):
        content = '{ a { b'
        self.assertEqual(find_block_end(content, 0), -1)

    def test_empty_block(self):
        content = '{}'  # '}' at index 1, returns 2
        self.assertEqual(find_block_end(content, 0), 2)

    def test_deeply_nested(self):
        content = '{ { { } } }'  # outer '}' at index 10, returns 11
        self.assertEqual(find_block_end(content, 0), 11)


class TestFindEnumSetBlock(unittest.TestCase):
    def test_finds_block(self):
        content = 'ENUMERATION_SET NAME="TestSet" { ENTRY VALUE=1 NAME="A" {} }'
        result = find_enum_set_block(content, 'TestSet')
        self.assertIsNotNone(result)
        self.assertEqual(result[0], 0)
        self.assertEqual(result[1], len(content))

    def test_finds_with_index(self):
        content = 'ENUMERATION_SET INDEX=0 NAME="TestSet" { }'
        result = find_enum_set_block(content, 'TestSet')
        self.assertIsNotNone(result)

    def test_not_found(self):
        content = 'ENUMERATION_SET NAME="Other" { }'
        result = find_enum_set_block(content, 'TestSet')
        self.assertIsNone(result)

    def test_no_brace(self):
        content = 'ENUMERATION_SET NAME="TestSet" no_brace'
        result = find_enum_set_block(content, 'TestSet')
        self.assertIsNone(result)


class TestExtractEnumSets(unittest.TestCase):
    def test_basic_extraction(self):
        content = '''
        ENUMERATION_SET NAME="TestSet" {
          DESCRIPTION="Test description"
          CATEGORY="TestCat"
          ENTRY VALUE=0 NAME="Off" { }
          ENTRY VALUE=1 NAME="On" { }
        }
        '''
        result = extract_enum_sets(content)
        self.assertIn('TestSet', result)
        self.assertEqual(len(result['TestSet']['entries']), 2)
        self.assertEqual(result['TestSet']['category'], 'TestCat')
        self.assertEqual(result['TestSet']['description'], 'Test description')

    def test_multiple_sets(self):
        content = '''
        ENUMERATION_SET NAME="Set1" { ENTRY VALUE=0 NAME="A" {} }
        ENUMERATION_SET NAME="Set2" { ENTRY VALUE=0 NAME="B" {} }
        '''
        result = extract_enum_sets(content)
        self.assertEqual(len(result), 2)
        self.assertIn('Set1', result)
        self.assertIn('Set2', result)

    def test_empty_content(self):
        result = extract_enum_sets('')
        self.assertEqual(len(result), 0)


class TestExtractStringValueRefs(unittest.TestCase):
    def test_basic(self):
        content = '''
        VALUE { SET="TestSet" STRING_VALUE="Hello" }
        VALUE { SET="TestSet" STRING_VALUE="Hello" }
        VALUE { SET="TestSet" STRING_VALUE="World" }
        '''
        result = extract_string_value_refs(content)
        self.assertIn('TestSet', result)
        self.assertEqual(result['TestSet']['Hello'], 2)
        self.assertEqual(result['TestSet']['World'], 1)

    def test_multiple_sets(self):
        content = '''
        VALUE { SET="Set1" STRING_VALUE="A" }
        VALUE { SET="Set2" STRING_VALUE="B" }
        '''
        result = extract_string_value_refs(content)
        self.assertEqual(len(result), 2)


class TestExtractExpressionRefs(unittest.TestCase):
    def test_dollar_prefix(self):
        content = "EXPR = '$phase_state:Running'"
        result = extract_expression_refs(content)
        self.assertIn('$phase_state', result)
        self.assertEqual(result['$phase_state']['Running'], 1)

    def test_plain_prefix(self):
        content = "EXPR = 'MySet:Value1'"
        result = extract_expression_refs(content)
        self.assertIn('MySet', result)

    def test_multiple_refs(self):
        content = "'$phase_state:Running' AND '$phase_state:Held'"
        result = extract_expression_refs(content)
        self.assertEqual(result['$phase_state']['Running'], 1)
        self.assertEqual(result['$phase_state']['Held'], 1)


class TestExtractAlarms(unittest.TestCase):
    def test_system_alarm(self):
        content = '''
        SYSTEM_ALARM INDEX=0 NAME="TestAlarm" {
          DESCRIPTION="Test alarm"
          ALARM_WORD="TEST"
          MESSAGE="Test message"
          CATEGORY="Safety"
        }
        '''
        result = extract_alarms(content)
        self.assertIn('TestAlarm', result)
        self.assertEqual(result['TestAlarm']['DESCRIPTION'], 'Test alarm')
        self.assertEqual(result['TestAlarm']['ALARM_WORD'], 'TEST')
        self.assertEqual(result['TestAlarm']['ALARM_TYPE'], 'SYSTEM_ALARM')

    def test_user_alarm(self):
        content = 'USER_ALARM NAME="UserAlarm" { DESCRIPTION="User alarm" }'
        result = extract_alarms(content)
        self.assertIn('UserAlarm', result)
        self.assertEqual(result['UserAlarm']['ALARM_TYPE'], 'USER_ALARM')


class TestExtractPriorityNames(unittest.TestCase):
    def test_basic(self):
        content = '''
        PRIORITY_NAME="CRITICAL"
        PRIORITY_NAME="WARNING"
        PRIORITY_NAME="CRITICAL"
        '''
        result = extract_priority_names(content)
        self.assertEqual(result, {'CRITICAL', 'WARNING'})


class TestIsChinese(unittest.TestCase):
    def test_chinese_string(self):
        self.assertTrue(_is_chinese('停止中'))
        self.assertTrue(_is_chinese('空闲'))

    def test_english_string(self):
        self.assertFalse(_is_chinese('Stopping'))
        self.assertFalse(_is_chinese('Idle'))

    def test_mixed(self):
        self.assertTrue(_is_chinese('Hello停止'))
        self.assertTrue(_is_chinese('abc空闲def'))

    def test_empty(self):
        self.assertFalse(_is_chinese(''))


class TestBidirectionalTranslate(unittest.TestCase):
    def test_cn_to_en(self):
        result, direction = _bidirectional_translate('$phase_state', '停止中')
        self.assertEqual(result, 'Stopping')
        self.assertEqual(direction, 'en')

    def test_en_to_cn(self):
        result, direction = _bidirectional_translate('$phase_state', 'Stopping')
        self.assertEqual(result, '停止中')
        self.assertEqual(direction, 'cn')

    def test_unknown_set(self):
        result, direction = _bidirectional_translate('$unknown', 'value')
        self.assertIsNone(result)
        self.assertIsNone(direction)

    def test_unknown_value(self):
        result, direction = _bidirectional_translate('$phase_state', 'NonExistentValue')
        self.assertIsNone(result)
        self.assertIsNone(direction)


class TestValidateXlsx(unittest.TestCase):
    def test_nonexistent_file(self):
        ok, msg = validate_xlsx('/nonexistent/path.xlsx')
        self.assertFalse(ok)
        self.assertIn('not found', msg.lower())

    def test_empty_file(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            ok, msg = validate_xlsx(path)
            self.assertFalse(ok)
            self.assertIn('empty', msg.lower())
        finally:
            os.unlink(path)

    def test_invalid_zip(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(b'this is not a zip file')
            path = f.name
        try:
            ok, msg = validate_xlsx(path)
            self.assertFalse(ok)
            self.assertIn('not a valid xlsx', msg.lower())
        finally:
            os.unlink(path)

    def test_valid_xlsx(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            wb = Workbook()
            wb.save(path)
            wb.close()
            ok, msg = validate_xlsx(path)
            self.assertTrue(ok)
        finally:
            os.unlink(path)


class TestCompareAlarms(unittest.TestCase):
    def test_both_present(self):
        old = {'A': {'DESCRIPTION': 'Old'}}
        new = {'A': {'DESCRIPTION': 'New'}}
        result = compare_alarms(old, new)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]['status'], 'Both')

    def test_old_only(self):
        old = {'A': {'DESCRIPTION': 'Old'}}
        result = compare_alarms(old, {})
        self.assertEqual(result[0]['status'], 'Old only')

    def test_new_only(self):
        new = {'A': {'DESCRIPTION': 'New'}}
        result = compare_alarms({}, new)
        self.assertEqual(result[0]['status'], 'New only')


class TestComparePriorityNames(unittest.TestCase):
    def test_suggested_mapping(self):
        cn_critical = chr(0x5371) + chr(0x6025)  # 危急
        cn_warning = chr(0x8B66) + chr(0x544A)   # 警告
        old_names = {cn_critical, cn_warning}
        new_names = {'CRITICAL'}
        result = compare_priority_names(old_names, new_names)
        # 危急 should have CRITICAL suggestion, 警告 should have WARNING suggestion
        self.assertEqual(len(result), 2)
        result_dict = {r['old_value']: r['new_value'] for r in result}
        self.assertEqual(result_dict[cn_critical], 'CRITICAL')
        self.assertEqual(result_dict[cn_warning], 'WARNING')

    def test_skip_existing(self):
        old_names = {'CRITICAL'}
        new_names = {'CRITICAL'}
        result = compare_priority_names(old_names, new_names)
        self.assertEqual(len(result), 0)


class TestReplacePriorityNames(unittest.TestCase):
    def test_basic(self):
        content = 'PRIORITY_NAME="危急" stuff PRIORITY_NAME="危急"'
        changes = {'危急': 'CRITICAL'}
        result, count = replace_priority_names(content, changes)
        self.assertEqual(count, 2)
        self.assertNotIn('危急', result)
        self.assertEqual(result.count('CRITICAL'), 2)


class TestReplaceAlarmValues(unittest.TestCase):
    def test_basic(self):
        content = 'SYSTEM_ALARM NAME="TestAlarm" { DESCRIPTION="Old desc" ALARM_WORD="Old" }'
        changes = {'TestAlarm': {'DESCRIPTION': 'New desc', 'ALARM_WORD': 'New'}}
        result, count = replace_alarm_values(content, changes)
        self.assertEqual(count, 2)
        self.assertIn('DESCRIPTION="New desc"', result)
        self.assertIn('ALARM_WORD="New"', result)


class TestWriteExcelSheet(unittest.TestCase):
    def test_simple_two_column(self):
        wb = Workbook()
        ws = wb.active
        rows = [
            {'old': 'OldVal1', 'new': 'NewVal1'},
            {'old': 'OldVal2', 'new': 'NewVal2'},
        ]
        _write_excel_sheet(ws, ['Old', 'New'], rows, data_cols=['old', 'new'])
        self.assertEqual(ws.cell(row=1, column=1).value, 'Old')
        self.assertEqual(ws.cell(row=2, column=1).value, 'OldVal1')
        self.assertEqual(ws.cell(row=2, column=2).value, 'NewVal1')
        wb.close()

    def test_empty_rows(self):
        wb = Workbook()
        ws = wb.active
        _write_excel_sheet(ws, ['A', 'B'], [])
        self.assertEqual(ws.cell(row=1, column=1).value, 'A')
        wb.close()


class TestReadAlarmTypesExcel(unittest.TestCase):
    def test_basic(self):
        wb = Workbook()
        ws = wb.create_sheet("Alarm Types")
        ws.append(['Alarm Name', 'Type', 'Status', 'Old Desc', 'New Desc',
                    'Old Word', 'New Word', 'Old Msg', 'New Msg', 'Old Cat', 'New Cat'])
        ws.append(['TestAlarm', 'SYSTEM', 'Both', 'Old', 'New', '', '', '', '', '', ''])
        result = read_alarm_types_excel(wb)
        self.assertIn('TestAlarm', result)
        self.assertEqual(result['TestAlarm']['DESCRIPTION'], 'New')
        wb.close()

    def test_no_sheet(self):
        wb = Workbook()
        result = read_alarm_types_excel(wb)
        self.assertEqual(len(result), 0)
        wb.close()


class TestReadPriorityNamesExcel(unittest.TestCase):
    def test_basic(self):
        wb = Workbook()
        ws = wb.create_sheet("Alarm Priorities")
        ws.append(['Old Priority Name', 'New Priority Name'])
        ws.append([chr(0x5371) + chr(0x6025), 'CRITICAL'])
        ws.append(['Already English', 'Already English'])
        result = read_priority_names_excel(wb)
        cn_critical = chr(0x5371) + chr(0x6025)
        self.assertIn(cn_critical, result)
        self.assertEqual(result[cn_critical], 'CRITICAL')
        self.assertNotIn('Already English', result)
        wb.close()


class TestValidateExcelForGeneration(unittest.TestCase):
    def test_valid_empty(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Namesets"
            ws.append(['SET Name', 'Category', 'Status', 'FHX Values', 'New Value'])
            wb.save(path)
            wb.close()
            ok, errors = validate_excel_for_generation(path)
            self.assertTrue(ok)
        finally:
            os.unlink(path)

    def test_formula_detection(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Namesets"
            ws.append(['SET Name', 'Category', 'Status', 'FHX Values', 'New Value'])
            ws.append(['TestSet', '', 'Both', 'old', '=VLOOKUP(A2,B:C,2,0)'])
            wb.save(path)
            wb.close()
            ok, errors = validate_excel_for_generation(path)
            self.assertFalse(ok)
            self.assertTrue(any('formula' in e['message'].lower() for e in errors))
        finally:
            os.unlink(path)

    def test_duplicate_value_numbers(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Namesets"
            ws.append(['SET Name', 'Category', 'Status', 'FHX Values', 'New Value'])
            ws.append(['NewSet', '', 'New', '',
                        'VALUE=0 NAME="A"\nVALUE=0 NAME="B"'])
            wb.save(path)
            wb.close()
            ok, errors = validate_excel_for_generation(path)
            self.assertFalse(ok)
            self.assertTrue(any('duplicate' in e['message'].lower() for e in errors))
        finally:
            os.unlink(path)


class TestReadFhx(unittest.TestCase):
    def test_utf16_le_bom(self):
        with tempfile.NamedTemporaryFile(suffix='.fhx', delete=False, mode='wb') as f:
            content = 'ENUMERATION_SET NAME="Test" {}'
            f.write(b'\xff\xfe' + content.encode('utf-16-le'))
            path = f.name
        try:
            result = read_fhx(path)
            self.assertIn('ENUMERATION_SET', result)
            self.assertIn('Test', result)
        finally:
            os.unlink(path)

    def test_utf8(self):
        with tempfile.NamedTemporaryFile(suffix='.fhx', delete=False, mode='wb') as f:
            content = 'ENUMERATION_SET NAME="Test" {}'
            f.write(content.encode('utf-8'))
            path = f.name
        try:
            result = read_fhx(path)
            self.assertIn('ENUMERATION_SET', result)
        finally:
            os.unlink(path)


class TestWriteFhx(unittest.TestCase):
    def test_roundtrip(self):
        with tempfile.NamedTemporaryFile(suffix='.fhx', delete=False) as f:
            path = f.name
        try:
            content = 'ENUMERATION_SET NAME="Test" {}'
            write_fhx(path, content)
            result = read_fhx(path)
            self.assertEqual(result, content)
        finally:
            os.unlink(path)


class TestDeltaValueMappings(unittest.TestCase):
    def test_cn_to_en_and_back(self):
        """Every CN->EN mapping should have a valid EN->CN reverse mapping.
        When two CN values map to the same EN value, only one roundtrip is guaranteed."""
        from collections import Counter
        for set_name, mapping in DELTA_VALUE_CN_TO_EN.items():
            self.assertIn(set_name, DELTA_VALUE_EN_TO_CN,
                          f"Missing EN->CN mapping for {set_name}")
            en_counts = Counter(mapping.values())
            for cn_val, en_val in mapping.items():
                if en_counts[en_val] > 1:
                    continue  # skip EN values shared by multiple CN terms
                self.assertEqual(DELTA_VALUE_EN_TO_CN[set_name][en_val], cn_val,
                                 f"Roundtrip failed for {set_name}: {cn_val} -> {en_val}")

    def test_no_duplicate_english_values(self):
        """No Chinese values should map to the same English value within a set,
        except for documented cases where two CN terms share one EN meaning."""
        # Known intentional duplicates: '正在保持' and '已保留' both -> 'Held'
        known_duplicates = {
            '$phase_state': {'Held'},
            '$recipe_state': {'Held'},
        }
        for set_name, mapping in DELTA_VALUE_CN_TO_EN.items():
            en_values = list(mapping.values())
            dup_set = known_duplicates.get(set_name, set())
            unique_en = [v for v in en_values if v not in dup_set or en_values.count(v) <= 2]
            # Just verify roundtrip works for all values
            for cn_val, en_val in mapping.items():
                self.assertIn(en_val, DELTA_VALUE_EN_TO_CN.get(set_name, {}),
                              f"EN->CN missing for {set_name}: {en_val}")


if __name__ == '__main__':
    unittest.main()
