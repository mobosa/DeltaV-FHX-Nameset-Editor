"""
DeltaV FHX Nameset Editor
Translates nameset values between Chinese and English in DeltaV FHX configuration files.
Handles any FHX type (Library, Control Strategies, Setup, Recipes, etc.) automatically.
Workflow:
  Step 1: Load FHX + Setup → Compare all nameset values → Export Excel
  Step 2: Import edited Excel → Generate new FHX with translated values
"""

import re
import os
import sys
import threading
import zipfile
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def validate_xlsx(filepath):
    """Validate that a file is a valid xlsx (ZIP) file. Returns (ok, error_msg)."""
    if not os.path.exists(filepath):
        return False, f"File not found: {filepath}"
    size = os.path.getsize(filepath)
    if size == 0:
        return False, f"File is empty (0 bytes): {filepath}"
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            bad = zf.testzip()
            if bad is not None:
                return False, f"Corrupt entry in xlsx: {bad}"
        return True, ''
    except zipfile.BadZipFile:
        return False, f"File is not a valid xlsx (not a ZIP file, {size} bytes): {filepath}"

def _safe_load_workbook(excel_path):
    """Load an Excel file with retry and fallback. Handles Excel file locking and format issues."""
    import time
    last_err = None
    for attempt in range(3):
        try:
            return load_workbook(excel_path, read_only=True)
        except Exception as e:
            last_err = e
            if 'not a zip' in str(e).lower() or 'badzip' in str(e).lower().__class__.__name__:
                # File might still be locked by Excel, wait and retry
                time.sleep(1)
                continue
            raise
    # All retries failed with zip error - try without read_only
    try:
        return load_workbook(excel_path, read_only=False)
    except Exception:
        pass
    # Last resort: try to read as data_only
    try:
        return load_workbook(excel_path, read_only=True, data_only=True)
    except Exception:
        pass
    raise RuntimeError(
        f"Cannot read Excel file. Please make sure:\n"
        f"1. The file is closed in Excel\n"
        f"2. The file is saved as .xlsx format (not .xls)\n"
        f"3. The file is not corrupted\n\n"
        f"Original error: {last_err}"
    )

# ============================================================
# File I/O
# ============================================================
def read_fhx(filepath):
    """Read FHX file (UTF-16 LE with BOM or UTF-8)."""
    with open(filepath, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        return raw[2:].decode('utf-16-le')
    if raw[:2] == b'\xfe\xff':
        return raw[2:].decode('utf-16-be')
    try:
        return raw.decode('utf-8')
    except:
        return raw.decode('utf-16-le', errors='replace')

def write_fhx(filepath, content):
    """Write FHX file as UTF-16 LE with BOM."""
    with open(filepath, 'wb') as f:
        f.write(b'\xff\xfe')
        f.write(content.encode('utf-16-le'))

# ============================================================
# Extract namesets from content
# ============================================================
def extract_enum_sets(content):
    """Extract all ENUMERATION_SET definitions. Returns dict: name -> {'entries': [...], 'category': '...', 'description': '...'}."""
    enum_sets = {}
    for m in re.finditer(r'ENUMERATION_SET\s+(?:INDEX=\d+\s+)?NAME="([^"]*)"', content):
        name = m.group(1)
        start = m.start()
        brace_start = content.find('{', start)
        if brace_start < 0:
            continue
        depth = 1
        pos = brace_start + 1
        while pos < len(content) and depth > 0:
            if content[pos] == '{':
                depth += 1
            elif content[pos] == '}':
                depth -= 1
            pos += 1
        if depth == 0:
            block = content[start:pos]
            values = []
            for em in re.finditer(r'ENTRY\s+VALUE=(\d+)\s+NAME="([^"]*)"', block):
                entry_str = f'VALUE={em.group(1)} NAME="{em.group(2)}"'
                values.append(entry_str)
            cat_match = re.search(r'CATEGORY="([^"]*)"', block)
            category = cat_match.group(1) if cat_match else ''
            desc_match = re.search(r'DESCRIPTION="([^"]*)"', block)
            description = desc_match.group(1) if desc_match else ''
            enum_sets[name] = {'entries': values, 'category': category, 'description': description}
    return enum_sets

# ============================================================
# Extraction helpers
# ============================================================
def extract_string_value_refs(content):
    """Extract SET+STRING_VALUE pairs from FHX. Returns dict: set_name -> {value -> count}."""
    refs = {}
    # Match SET and STRING_VALUE on separate lines within VALUE blocks
    for m in re.finditer(r'SET="([^"]*)"[^}]*?STRING_VALUE="([^"]*)"', content, re.DOTALL):
        set_name = m.group(1)
        val = m.group(2)
        if set_name not in refs:
            refs[set_name] = {}
        refs[set_name][val] = refs[set_name].get(val, 0) + 1
    return refs

def extract_expression_refs(content):
    """Extract nameset references from expressions. Returns dict: set_name -> {value -> count}.
    Matches patterns like '$nameset:value', '_nameset:value', or 'nameset:value' in expressions.
    Nameset names can contain letters, digits, underscores, and hyphens."""
    refs = {}
    # Match 'nameset_name:value' pattern (supports $prefix, _prefix, or plain names, with hyphens)
    # Only match when nameset name starts with letter/underscore/$ and value is a simple identifier
    for m in re.finditer(r"'(\$?[a-zA-Z_][a-zA-Z0-9_-]*):([^']+)'", content):
        set_name = m.group(1)
        val = m.group(2).strip()
        if not val:
            continue
        if set_name not in refs:
            refs[set_name] = {}
        refs[set_name][val] = refs[set_name].get(val, 0) + 1
    return refs

def extract_alarms(content):
    """Extract SYSTEM_ALARM and USER_ALARM blocks. Returns dict: alarm_name -> dict of fields."""
    alarms = {}
    for m in re.finditer(r'(SYSTEM_ALARM|USER_ALARM)\s+(?:INDEX=(\d+)\s+)?NAME="([^"]*)"', content):
        alarm_type = m.group(1)
        alarm_index = m.group(2)
        alarm_name = m.group(3)
        start = m.start()
        brace_start = content.find('{', start)
        if brace_start < 0:
            continue
        depth = 1
        pos = brace_start + 1
        while pos < len(content) and depth > 0:
            if content[pos] == '{':
                depth += 1
            elif content[pos] == '}':
                depth -= 1
            pos += 1
        if depth == 0:
            block = content[start:pos]
            fields = {}
            for field in ('DESCRIPTION', 'ALARM_WORD', 'MESSAGE', 'CATEGORY', 'SUMMARY_NO',
                          'DEFAULT_PARAM1', 'DEFAULT_PARAM2', 'WAVE_FILE'):
                fm = re.search(rf'{field}="([^"]*)"', block)
                if fm:
                    fields[field] = fm.group(1)
            fields['ALARM_TYPE'] = alarm_type
            if alarm_index:
                fields['INDEX'] = alarm_index
            alarms[alarm_name] = fields
    return alarms

def compare_alarms(lib_or_cs_alarms, setup_alarms):
    """Compare alarm definitions. Returns list of dicts for Excel."""
    comparison = []
    all_names = sorted(set(list(lib_or_cs_alarms.keys()) + list(setup_alarms.keys())))
    for name in all_names:
        old_alarm = lib_or_cs_alarms.get(name, {})
        new_alarm = setup_alarms.get(name, {})
        if old_alarm and new_alarm:
            status = 'Both'
        elif old_alarm:
            status = 'Old only'
        else:
            status = 'New only'
        comparison.append({
            'name': name,
            'alarm_type': old_alarm.get('ALARM_TYPE', new_alarm.get('ALARM_TYPE', '')),
            'status': status,
            'old_description': old_alarm.get('DESCRIPTION', ''),
            'new_description': new_alarm.get('DESCRIPTION', ''),
            'old_alarm_word': old_alarm.get('ALARM_WORD', ''),
            'new_alarm_word': new_alarm.get('ALARM_WORD', ''),
            'old_message': old_alarm.get('MESSAGE', ''),
            'new_message': new_alarm.get('MESSAGE', ''),
            'old_category': old_alarm.get('CATEGORY', ''),
            'new_category': new_alarm.get('CATEGORY', ''),
        })
    return comparison

def write_alarm_types_sheet(wb, alarm_comparison, header_fill, header_font, thin_border,
                            yellow_fill, green_fill, red_fill):
    """Write the 'Alarm Types' sheet to an existing workbook."""
    if not alarm_comparison:
        return
    ws = wb.create_sheet("Alarm Types")
    headers = ['Alarm Name', 'Alarm Type', 'Status',
               'Old Description', 'New Description',
               'Old Alarm Word', 'New Alarm Word',
               'Old Message', 'New Message',
               'Old Category', 'New Category']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(alarm_comparison, 2):
        ws.cell(row=i, column=1, value=item['name']).border = thin_border
        ws.cell(row=i, column=2, value=item.get('alarm_type', '')).border = thin_border
        status_cell = ws.cell(row=i, column=3, value=item['status'])
        status_cell.border = thin_border
        if item['status'] == 'Both':
            status_cell.fill = green_fill
        elif item['status'] == 'Old only':
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = red_fill

        for col_idx, field in enumerate(['old_description', 'new_description',
                                          'old_alarm_word', 'new_alarm_word',
                                          'old_message', 'new_message',
                                          'old_category', 'new_category'], 4):
            cell = ws.cell(row=i, column=col_idx, value=item.get(field, ''))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:K{len(alarm_comparison) + 1}"

def read_alarm_types_excel(wb):
    """Read 'Alarm Types' sheet from workbook. Returns dict: alarm_name -> dict of new field values."""
    alarm_changes = {}
    if "Alarm Types" not in wb.sheetnames:
        return alarm_changes
    ws = wb["Alarm Types"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        new_description = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        new_alarm_word = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        new_message = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        new_category = str(row[10]).strip() if len(row) > 10 and row[10] else ''
        changes = {}
        if new_description:
            changes['DESCRIPTION'] = new_description
        if new_alarm_word:
            changes['ALARM_WORD'] = new_alarm_word
        if new_message:
            changes['MESSAGE'] = new_message
        if new_category:
            changes['CATEGORY'] = new_category
        if changes:
            alarm_changes[name] = changes
    return alarm_changes

def extract_priority_names(content):
    """Extract unique PRIORITY_NAME values from MODULE CLASS definitions. Returns set of values."""
    names = set()
    for m in re.finditer(r'PRIORITY_NAME="([^"]*)"', content):
        val = m.group(1).strip()
        if val:
            names.add(val)
    return names

def compare_priority_names(lib_or_cs_names, setup_names):
    """Compare PRIORITY_NAME values. Returns list of dicts for Excel."""
    # Build suggested mapping based on standard DeltaV priorities
    STANDARD_MAP = {
        chr(0x5371) + chr(0x6025): 'CRITICAL',   # 危急
        chr(0x8B66) + chr(0x544A): 'WARNING',    # 警告
        chr(0x63D0) + chr(0x793A): 'ADVISORY',   # 提示
        chr(0x8BB0) + chr(0x5F55): 'LOG',        # 记录
    }
    comparison = []
    for name in sorted(lib_or_cs_names):
        if name in setup_names:
            continue  # Already English, skip
        suggested = STANDARD_MAP.get(name, '')
        comparison.append({
            'old_value': name,
            'new_value': suggested,
            'count': content_count(name),
        })
    return comparison

def content_count(val):
    """Helper: returns 0 (actual count filled during export)."""
    return 0

def write_priority_names_sheet(wb, priority_comparison, header_fill, header_font, thin_border, yellow_fill):
    """Write 'Alarm Priorities' sheet to workbook."""
    if not priority_comparison:
        return
    ws = wb.create_sheet("Alarm Priorities")
    headers = ['Old Priority Name', 'New Priority Name']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(priority_comparison, 2):
        ws.cell(row=i, column=1, value=item['old_value']).border = thin_border
        new_cell = ws.cell(row=i, column=2, value=item['new_value'])
        new_cell.alignment = Alignment(wrap_text=True, vertical='top')
        new_cell.border = thin_border
        if not item['new_value']:
            new_cell.fill = yellow_fill

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:B{len(priority_comparison) + 1}"

def read_priority_names_excel(wb):
    """Read 'Alarm Priorities' sheet. Returns dict: old_value -> new_value."""
    changes = {}
    if "Alarm Priorities" not in wb.sheetnames:
        return changes
    ws = wb["Alarm Priorities"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        old_val = str(row[0]).strip() if row[0] else ''
        new_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if old_val and new_val and old_val != new_val:
            changes[old_val] = new_val
    return changes

def replace_priority_names(content, priority_changes):
    """Replace PRIORITY_NAME values in MODULE CLASS definitions."""
    count = 0
    for old_val, new_val in priority_changes.items():
        old_pattern = f'PRIORITY_NAME="{old_val}"'
        new_pattern = f'PRIORITY_NAME="{new_val}"'
        n = content.count(old_pattern)
        content = content.replace(old_pattern, new_pattern)
        count += n
    return content, count

def replace_alarm_values(content, alarm_changes):
    """Replace DESCRIPTION, ALARM_WORD, MESSAGE, CATEGORY in SYSTEM_ALARM/USER_ALARM blocks."""
    replace_count = 0
    for alarm_name, changes in alarm_changes.items():
        escaped_name = re.escape(alarm_name)
        pattern = r'(SYSTEM_ALARM|USER_ALARM)\s+(?:INDEX=\d+\s+)?NAME="' + escaped_name + r'"'
        m = re.search(pattern, content)
        if not m:
            continue
        alarm_type = m.group(1)
        start = m.start()
        brace_pos = content.find('{', start)
        if brace_pos < 0:
            continue
        depth = 1
        pos = brace_pos + 1
        while pos < len(content) and depth > 0:
            if content[pos] == '{':
                depth += 1
            elif content[pos] == '}':
                depth -= 1
            pos += 1
        if depth != 0:
            continue
        block = content[start:pos]
        for field, new_val in changes.items():
            fm = re.search(rf'{field}="[^"]*"', block)
            if fm:
                old_val = fm.group(0)
                new_field = f'{field}="{new_val}"'
                block = block.replace(old_val, new_field, 1)
                replace_count += 1
        content = content[:start] + block + content[pos:]
    return content, replace_count

# Hardcoded Chinese→English mapping for standard DeltaV nameset values
# These namesets are system-defined (not in FHX ENUMERATION_SET) so VALUE number
# cross-referencing is not possible. Mapping derived from DeltaV standard definitions.
DELTA_VALUE_CN_TO_EN = {
    '$phase_state': {
        '停止中': 'Stopping', '正在退出': 'Aborting', '空闲': 'Idle',
        '正在运行': 'Running', '已退出': 'Aborted', '已完成': 'Complete',
        '保留中': 'Holding', '正在启动': 'Starting', '正在保持': 'Held',
        '已保留': 'Held', '准备就绪': 'Ready', '已停止': 'Stopped',
        '正在重启': 'Restarting', '未加载': 'Not Loaded',
    },
    '$recipe_state': {
        '已完成': 'Complete', '停止中': 'Stopping', '正在退出': 'Aborting',
        '空闲': 'Idle', '正在运行': 'Running', '已退出': 'Aborted',
        '保留中': 'Holding', '正在启动': 'Starting', '正在保持': 'Held',
        '已保留': 'Held', '准备就绪': 'Ready', '已停止': 'Stopped',
        '正在重启': 'Restarting', '未加载': 'Not Loaded',
    },
    '$sfc_action_states': {
        '已完成': 'Complete', '激活': 'Active', '不活动': 'Inactive',
        '延迟': 'Delayed', '待定': 'Pending', '失败': 'Failed',
    },
    '$phase_command': {
        '清除故障': 'Clear Failures', '中止': 'Abort', '保持': 'Hold',
        '停止': 'Stop', '复位': 'Reset', '暂停': 'Pause',
        '下载': 'Download', '恢复': 'Resume', '重启': 'Restart',
        '启动': 'Start', '自动': 'Automatic', '手动': 'Manual',
    },
    '$phase_owner_id': {
        'DeltaV 批量': 'DeltaV Batch', '外部': 'External',
    },
    '$sfc_commands': {
        '启动顺控': 'Start Sequence', '停止顺控': 'Stop Sequence',
        '复位顺控': 'Reset Sequence',
    },
    '$sfc_states': {
        '顺控空闲': 'Sequence Idle', '顺控激活': 'Sequence Active',
        '顺控停止': 'Sequence Stopped', '顺控完成': 'Sequence Completed',
        '顺控堵塞': 'Sequence Blocked',
    },
    '$phase_restart_types': {
        '继续': 'Continue', '从启动开始': 'From Start',
        '从启动开始无下载': 'From Start Without Download',
    },
    '$phase_wdog_states': {
        '已失败': 'FAILED', '可疑': 'SUSPECT', '正常': 'OK',
    },
    '$dc_states': {
        '锁定': 'Locked', '已跳车': 'Tripped',
        '非励磁回讯失败': 'Failed Passive', '停车/联锁': 'Shutdown/Interlocked',
    },
    '$sysstat_opts': {
        '切换': 'Switchover', '我的下装': 'MyDownload',
        '全部下装': 'TotalDownload', '电源故障': 'Powerfail',
    },
    '$time_format': {
        '本地': 'Local', '协调世界时': 'UTC',
    },
    '$module_states': {
        '离线': 'Out Of Service', '服务中': 'In Service',
    },
    'Alarm_Group_Mode': {
        '已跳车': 'Tripped', '正常': 'Normal', '已停用': 'Disabled',
    },
    'Suppress_Timeout_Opt': {
        '所有已抑制': 'All Suppressed', '激活已抑制': 'Active Suppressed',
        '无抑制': 'None Suppressed',
    },
    'LOOP_TYPE': {
        '自动': 'Auto', '手动': 'Manual', '结束': 'End', '继续': 'Continue',
    },
    'MAN_RETURN': {
        '自动': 'Auto', '自动回手动': 'Auto To Manual',
        '继续': 'Continue', '重新输入': 'Re-enter',
    },
    'RESTART_MODE': {
        '启动': 'Start', '无等待启动': 'No Wait Start',
        '开始': 'Start', '继续等待开始': 'Continue Wait Start',
    },
    '$isel_typ': {
        '第一个有效值': 'First Valid', '最小': 'Minimum', '最大': 'Maximum',
        '平均': 'Average',
    },
    '$l_typ': {
        '间接': 'Indirect', '直接': 'Direct',
    },
    '$sysstat_ls_opts': {
        '电源故障': 'Powerfail',
    },
    '$time': {
        '秒': 'Seconds', '分': 'Minutes', '时': 'Hours',
    },
}

# ============================================================
# FHX Migration
# ============================================================
DELTA_VALUE_EN_TO_CN = {}
for _set, _map in DELTA_VALUE_CN_TO_EN.items():
    for _cn, _en in _map.items():
        DELTA_VALUE_EN_TO_CN.setdefault(_set, {})[_en] = _cn
del _set, _map, _cn, _en


def _bidirectional_translate(set_name, value):
    """Translate a nameset value bidirectionally based on content.
    Chinese value → English; English value → Chinese."""
    if re.search(r'[一-鿿]', value):
        # Value is Chinese → translate to English
        if set_name in DELTA_VALUE_CN_TO_EN and value in DELTA_VALUE_CN_TO_EN[set_name]:
            return DELTA_VALUE_CN_TO_EN[set_name][value], 'en'
    else:
        # Value is not Chinese (likely English) → translate to Chinese
        if set_name in DELTA_VALUE_EN_TO_CN and value in DELTA_VALUE_EN_TO_CN[set_name]:
            return DELTA_VALUE_EN_TO_CN[set_name][value], 'cn'
    return None, None


def compare_lib_and_export(lib_path, setup_path, output_path, log_callback=None, progress_callback=None):
    """Compare FHX ENUMERATION_SET + STRING_VALUE + Expression refs with Setup. Export Excel with three sheets."""
    def log(msg):
        if log_callback:
            log_callback(msg)

    def progress(pct, text=''):
        if progress_callback:
            progress_callback(pct, text)

    log(f"Reading FHX: {lib_path}")
    progress(0, 'Reading FHX...')
    lib_content = read_fhx(lib_path)
    log(f"  {len(lib_content):,} chars")

    log(f"Reading Setup: {setup_path}")
    progress(15, 'Reading Setup...')
    setup_content = read_fhx(setup_path)
    log(f"  {len(setup_content):,} chars")

    # Extract ENUMERATION_SET definitions
    progress(25, 'Extracting ENUMERATION_SET definitions...')
    lib_enum_defs = extract_enum_sets(lib_content)
    setup_enum_defs = extract_enum_sets(setup_content)
    log(f"  FHX ENUMERATION_SET definitions: {len(lib_enum_defs)}")
    log(f"  Setup ENUMERATION_SET definitions: {len(setup_enum_defs)}")

    # Extract STRING_VALUE references
    progress(35, 'Extracting STRING_VALUE references...')
    lib_refs = extract_string_value_refs(lib_content)
    log(f"  FHX SET+STRING_VALUE refs: {len(lib_refs)}")

    # Extract expression references
    progress(45, 'Extracting expression references...')
    expr_refs = extract_expression_refs(lib_content)
    log(f"  FHX expression refs: {len(expr_refs)}")

    # Build ENUMERATION_SET comparison
    progress(50, 'Building ENUMERATION_SET comparison...')
    nameset_comparison = []
    for set_name in sorted(lib_enum_defs.keys()):
        lib_data = lib_enum_defs[set_name]
        lib_vals = lib_data['entries'] if isinstance(lib_data, dict) else lib_data
        lib_cat = lib_data['category'] if isinstance(lib_data, dict) else ''
        lib_desc = lib_data['description'] if isinstance(lib_data, dict) else ''

        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'category': '', 'description': ''})
        setup_vals = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_cat = setup_data['category'] if isinstance(setup_data, dict) else ''

        if set_name in setup_enum_defs:
            status = 'Both'
            default_suggestion = '\n'.join(setup_vals) if setup_vals else ''
            category = setup_cat
        else:
            status = 'FHX only'
            default_suggestion = ''
            category = lib_cat

        nameset_comparison.append({
            'name': set_name,
            'fhx_values': lib_vals,
            'status': status,
            'default_suggestion': default_suggestion,
            'category': category,
            'description': lib_desc,
        })

    # Build STRING_VALUE comparison
    progress(60, 'Building STRING_VALUE comparison...')
    sv_comparison = []
    for set_name in sorted(lib_refs.keys()):
        vals = lib_refs[set_name]
        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'description': ''})
        setup_entries = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_desc = setup_data['description'] if isinstance(setup_data, dict) else ''

        entry_map = {}
        for entry in setup_entries:
            em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
            if em:
                entry_map[em.group(2)] = entry

        for val in sorted(vals.keys()):
            count = vals[val]
            suggested = ''
            for entry_name, entry_str in entry_map.items():
                if val == entry_name:
                    suggested = entry_str
                    break
            if not suggested:
                val_m = re.match(r'VALUE=(\d+)', val)
                if val_m:
                    for entry_name, entry_str in entry_map.items():
                        if f'VALUE={val_m.group(1)}' in entry_str:
                            suggested = entry_str
                            break
            # Fallback: bidirectional translation mapping for standard DeltaV namesets
            if not suggested and set_name in DELTA_VALUE_CN_TO_EN:
                trans_val, direction = _bidirectional_translate(set_name, val)
                if trans_val is not None and trans_val in entry_map:
                    suggested = entry_map[trans_val]

            sv_comparison.append({
                'set_name': set_name,
                'current_value': val,
                'count': count,
                'suggested': suggested,
                'description': setup_desc,
                'setup_entries': '\n'.join(setup_entries),
            })

    # Build expression comparison
    progress(70, 'Building expression comparison...')
    expr_comparison = []
    for set_name in sorted(expr_refs.keys()):
        vals = expr_refs[set_name]
        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'description': ''})
        setup_entries = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_desc = setup_data['description'] if isinstance(setup_data, dict) else ''

        entry_map = {}
        for entry in setup_entries:
            em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
            if em:
                entry_map[em.group(2)] = entry

        # Build Library entry map for fallback (for custom namesets not in Setup)
        lib_entry_map = {}
        lib_name_to_num = {}
        lib_num_to_entry = {}
        if set_name in lib_enum_defs:
            for entry in lib_enum_defs[set_name]['entries']:
                em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
                if em:
                    lib_entry_map[em.group(2)] = entry
                    lib_name_to_num[em.group(2)] = em.group(1)
                    lib_num_to_entry[em.group(1)] = entry

        # Build Setup VALUE number map for cross-referencing
        setup_num_to_name = {}
        for entry_name, entry_str in entry_map.items():
            vm = re.match(r'VALUE=(\d+)', entry_str)
            if vm:
                setup_num_to_name[vm.group(1)] = entry_name

        for val in sorted(vals.keys()):
            count = vals[val]
            suggested = ''
            # Direct match by name in Setup
            for entry_name, entry_str in entry_map.items():
                if val == entry_name:
                    suggested = entry_str
                    break
            # Fallback: match Chinese value via Library VALUE number -> Setup English name
            if not suggested and val in lib_name_to_num:
                val_num = lib_name_to_num[val]
                if val_num in setup_num_to_name:
                    eng_name = setup_num_to_name[val_num]
                    suggested = f'VALUE={val_num} NAME="{eng_name}"'
            # Fallback: use Library ENUMERATION_SET entry if not in Setup
            if not suggested and val in lib_entry_map:
                suggested = lib_entry_map[val]
            # Fallback: bidirectional translation mapping for standard DeltaV namesets
            if not suggested and set_name in DELTA_VALUE_CN_TO_EN:
                trans_val, direction = _bidirectional_translate(set_name, val)
                if trans_val is not None:
                    # Find the translated value in Setup entry map
                    if trans_val in entry_map:
                        suggested = entry_map[trans_val]
                    elif setup_num_to_name:
                        for num, name in setup_num_to_name.items():
                            if name == trans_val:
                                suggested = f'VALUE={num} NAME="{trans_val}"'
                                break

            expr_comparison.append({
                'set_name': set_name,
                'current_value': val,
                'count': count,
                'suggested': suggested,
                'description': setup_desc,
                'setup_entries': '\n'.join(setup_entries) if setup_entries else '\n'.join(lib_enum_defs.get(set_name, {}).get('entries', [])),
            })

    # Extract and compare alarm types
    progress(78, 'Extracting alarm types...')
    lib_alarms = extract_alarms(lib_content)
    setup_alarms = extract_alarms(setup_content)
    alarm_comparison = compare_alarms(lib_alarms, setup_alarms)
    log(f"  FHX alarms: {len(lib_alarms)}, Setup alarms: {len(setup_alarms)}")

    # Extract PRIORITY_NAME values
    lib_priorities = extract_priority_names(lib_content)
    setup_priorities = extract_priority_names(setup_content)
    priority_comparison = compare_priority_names(lib_priorities, setup_priorities)
    for item in priority_comparison:
        item['count'] = lib_content.count(f'PRIORITY_NAME="{item["old_value"]}"')
    log(f"  FHX priorities: {len(lib_priorities)}, Setup priorities: {len(setup_priorities)}")

    # Write Excel with three sheets
    progress(80, 'Writing Excel...')
    write_lib_comparison_excel(nameset_comparison, sv_comparison, expr_comparison, output_path, alarm_comparison, priority_comparison)

    progress(100, 'Done')
    log(f"\nExcel exported: {output_path}")
    log(f"  ENUMERATION_SET definitions: {len(nameset_comparison)}")
    log(f"  Alarm types: {len(alarm_comparison)}")
    log(f"  STRING_VALUE references: {len(sv_comparison)}")
    log(f"  Expression references: {len(expr_comparison)}")

    # List nameset values line by line
    log("\n" + "=" * 60)
    log("ENUMERATION_SET DEFINITIONS:")
    log("=" * 60)
    for item in nameset_comparison:
        log(f"\n[{item['name']}] ({item['status']})")
        if item['fhx_values']:
            for val in item['fhx_values']:
                log(f"  Lib: {val}")
        if item['default_suggestion']:
            for val in item['default_suggestion'].split('\n'):
                if val.strip():
                    log(f"  Setup: {val.strip()}")

    log("\n" + "=" * 60)
    log("STRING_VALUE REFERENCES:")
    log("=" * 60)
    current_set = ''
    for item in sv_comparison:
        if item['set_name'] != current_set:
            current_set = item['set_name']
            log(f"\n[{current_set}]")
        count_str = f" (x{item['count']})" if item['count'] > 1 else ""
        suggested_str = f" -> {item['suggested']}" if item['suggested'] else " (no match)"
        log(f"  {item['current_value']}{count_str}{suggested_str}")

    log("\n" + "=" * 60)
    log("EXPRESSION REFERENCES:")
    log("=" * 60)
    current_set = ''
    for item in expr_comparison:
        if item['set_name'] != current_set:
            current_set = item['set_name']
            log(f"\n[{current_set}]")
        count_str = f" (x{item['count']})" if item['count'] > 1 else ""
        suggested_str = f" -> {item['suggested']}" if item['suggested'] else " (no match)"
        log(f"  {item['current_value']}{count_str}{suggested_str}")

    return nameset_comparison, sv_comparison, expr_comparison

def write_lib_comparison_excel(nameset_comparison, sv_comparison, expr_comparison, output_path, alarm_comparison=None, priority_comparison=None):
    """Write FHX comparison to Excel with three sheets."""
    wb = Workbook()

    # Sheet 1: ENUMERATION_SET definitions
    ws1 = wb.active
    ws1.title = "Namesets"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    light_blue_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

    headers = ['SET Name', 'Category', 'Status', 'FHX Values', 'New Value', 'Description']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(nameset_comparison, 2):
        ws1.cell(row=i, column=1, value=item['name']).border = thin_border
        ws1.cell(row=i, column=2, value=item.get('category', '')).border = thin_border

        status_cell = ws1.cell(row=i, column=3, value=item['status'])
        status_cell.border = thin_border
        if item['status'] == 'Both':
            status_cell.fill = green_fill
        else:
            status_cell.fill = red_fill

        fhx_vals_str = '\n'.join(item['fhx_values'])
        fhx_cell = ws1.cell(row=i, column=4, value=fhx_vals_str)
        fhx_cell.border = thin_border
        fhx_cell.alignment = Alignment(wrap_text=True, vertical='top')

        new_cell = ws1.cell(row=i, column=5, value=item['default_suggestion'])
        new_cell.alignment = Alignment(wrap_text=True, vertical='top')
        new_cell.border = thin_border
        if item['default_suggestion']:
            new_cell.fill = light_blue_fill
        else:
            new_cell.fill = yellow_fill

        desc_cell = ws1.cell(row=i, column=6, value=item.get('description', ''))
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        desc_cell.border = thin_border

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 40
    ws1.column_dimensions['F'].width = 40
    ws1.freeze_panes = 'A2'
    if nameset_comparison:
        ws1.auto_filter.ref = f"A1:F{len(nameset_comparison) + 1}"

    # Sheet 2: STRING_VALUE references
    ws2 = wb.create_sheet("String Values")

    headers2 = ['SET Name', 'Current Value', 'Count', 'New Value', 'Setup Entries', 'Description']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(sv_comparison, 2):
        ws2.cell(row=i, column=1, value=item['set_name']).border = thin_border

        curr_cell = ws2.cell(row=i, column=2, value=item['current_value'])
        curr_cell.border = thin_border
        curr_cell.alignment = Alignment(wrap_text=True, vertical='top')

        count_cell = ws2.cell(row=i, column=3, value=item['count'])
        count_cell.border = thin_border
        count_cell.alignment = Alignment(horizontal='center')

        new_cell = ws2.cell(row=i, column=4, value=item['suggested'])
        new_cell.alignment = Alignment(wrap_text=True, vertical='top')
        new_cell.border = thin_border
        if item['suggested']:
            new_cell.fill = light_blue_fill
        else:
            new_cell.fill = yellow_fill

        setup_cell = ws2.cell(row=i, column=5, value=item['setup_entries'])
        setup_cell.alignment = Alignment(wrap_text=True, vertical='top')
        setup_cell.border = thin_border

        desc_cell = ws2.cell(row=i, column=6, value=item.get('description', ''))
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        desc_cell.border = thin_border

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 40
    ws2.freeze_panes = 'A2'
    if sv_comparison:
        ws2.auto_filter.ref = f"A1:F{len(sv_comparison) + 1}"

    # Sheet 3: Expression references
    ws3 = wb.create_sheet("Expression Refs")

    headers3 = ['SET Name', 'Current Value', 'Count', 'New Value', 'Setup Entries', 'Description']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(expr_comparison, 2):
        ws3.cell(row=i, column=1, value=item['set_name']).border = thin_border

        curr_cell = ws3.cell(row=i, column=2, value=item['current_value'])
        curr_cell.border = thin_border
        curr_cell.alignment = Alignment(wrap_text=True, vertical='top')

        count_cell = ws3.cell(row=i, column=3, value=item['count'])
        count_cell.border = thin_border
        count_cell.alignment = Alignment(horizontal='center')

        new_cell = ws3.cell(row=i, column=4, value=item['suggested'])
        new_cell.alignment = Alignment(wrap_text=True, vertical='top')
        new_cell.border = thin_border
        if item['suggested']:
            new_cell.fill = light_blue_fill
        else:
            new_cell.fill = yellow_fill

        setup_cell = ws3.cell(row=i, column=5, value=item['setup_entries'])
        setup_cell.alignment = Alignment(wrap_text=True, vertical='top')
        setup_cell.border = thin_border

        desc_cell = ws3.cell(row=i, column=6, value=item.get('description', ''))
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        desc_cell.border = thin_border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 8
    ws3.column_dimensions['D'].width = 40
    ws3.column_dimensions['E'].width = 50
    ws3.column_dimensions['F'].width = 40
    ws3.freeze_panes = 'A2'
    if expr_comparison:
        ws3.auto_filter.ref = f"A1:F{len(expr_comparison) + 1}"

    write_alarm_types_sheet(wb, alarm_comparison, header_fill, header_font, thin_border,
                            yellow_fill, green_fill, red_fill)

    write_priority_names_sheet(wb, priority_comparison, header_fill, header_font, thin_border, yellow_fill)

    wb.save(output_path)
    ok, err = validate_xlsx(output_path)
    if not ok:
        raise RuntimeError(f"Failed to write valid Excel file: {err}")

def validate_excel_for_generation(excel_path, log_callback=None):
    """Validate Excel data before generating new FHX. Returns (is_valid, errors).

    Checks all sheets for common data issues that would cause problems during generation.
    Each error is a dict: {'sheet': str, 'row': int, 'column': str, 'message': str}.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    errors = []
    wb = _safe_load_workbook(excel_path)

    # --- Namesets sheet ---
    if "Namesets" in wb.sheetnames:
        ws = wb["Namesets"]
        seen_namesets = {}  # set_name -> {value_num: row}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 5:
                continue
            name, category, status, lib_vals_str, new_value = row[:5]
            if not name:
                continue
            set_name = str(name).strip()
            status_str = str(status).strip() if status else ''

            if new_value:
                nv_str = str(new_value).strip()
                # Check for Excel formulas (=VLOOKUP, =IF, =INDEX, etc.)
                if nv_str.startswith('='):
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'Cell contains a formula instead of plain text: "{nv_str[:80]}". '
                                   f'Formulas cannot be used - please replace with the actual value.',
                    })

            if status_str in ('', 'New'):
                # New nameset: must have valid entries in New Value
                entries_str = new_value if new_value else lib_vals_str
                if not entries_str:
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'New nameset "{set_name}" has no entries in New Value or FHX Values column',
                    })
                    continue
                entries_text = [v.strip() for v in str(entries_str).split('\n') if v.strip()]
                valid_count = 0
                seen_nums = {}
                for line_idx, line in enumerate(entries_text):
                    m = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', line)
                    if m:
                        val_num = int(m.group(1))
                        val_name = m.group(2)
                        if not val_name:
                            errors.append({
                                'sheet': 'Namesets', 'row': row_idx,
                                'column': 'New Value (E)',
                                'message': f'Entry has empty NAME: VALUE={val_num} NAME=""',
                            })
                        if val_num in seen_nums:
                            errors.append({
                                'sheet': 'Namesets', 'row': row_idx,
                                'column': 'New Value (E)',
                                'message': f'Duplicate VALUE number {val_num} in nameset "{set_name}" '
                                           f'(first at line {seen_nums[val_num]}, again at line {line_idx + 1})',
                            })
                        else:
                            seen_nums[val_num] = line_idx + 1
                        valid_count += 1
                    else:
                        errors.append({
                            'sheet': 'Namesets', 'row': row_idx,
                            'column': 'New Value (E)',
                            'message': f'Invalid entry format: "{line}" (expected VALUE=N NAME="...")',
                        })
                if valid_count == 0 and entries_text:
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'No valid VALUE=N NAME="..." entries found for new nameset "{set_name}"',
                    })
            else:
                # Existing nameset with changes: validate New Value format per line
                if new_value:
                    nv_str = str(new_value).strip()
                    nv_lines = [l.strip() for l in nv_str.split('\n') if l.strip()]
                    lib_lines = [l.strip() for l in str(lib_vals_str).split('\n') if l.strip()] if lib_vals_str else []
                    for line_idx, nv_line in enumerate(nv_lines):
                        # Valid formats: VALUE=N NAME="...", or a simple name (e.g. "STOP")
                        if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*")', nv_line):
                            # Could be a simple name - that's okay, will be auto-constructed
                            # But warn if it contains spaces or special chars that look wrong
                            if ' ' in nv_line and not nv_line.startswith('VALUE='):
                                errors.append({
                                    'sheet': 'Namesets', 'row': row_idx,
                                    'column': 'New Value (E)',
                                    'message': f'Dubious entry format (has spaces): "{nv_line}" - '
                                               f'Expected VALUE=N NAME="..." or simple name',
                                })

    # --- String Values sheet ---
    if "String Values" in wb.sheetnames:
        ws = wb["String Values"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if not new_value or new_value == current_value:
                continue

            # Validate new_value format
            if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*"|[^\s]+)', new_value):
                errors.append({
                    'sheet': 'String Values', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Invalid format: "{new_value}" for set "{set_name}", value "{current_value}"',
                })
            # Check for Excel formulas
            if new_value.startswith('='):
                errors.append({
                    'sheet': 'String Values', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Cell contains a formula instead of plain text: "{new_value[:80]}". '
                               f'Formulas cannot be used - please replace with the actual value.',
                })

    # --- Expression Refs sheet ---
    if "Expression Refs" in wb.sheetnames:
        ws = wb["Expression Refs"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if not new_value or new_value == current_value:
                continue

            # Validate new_value format
            if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*"|[^\s]+)', new_value):
                errors.append({
                    'sheet': 'Expression Refs', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Invalid format: "{new_value}" for set "{set_name}", value "{current_value}"',
                })
            # Check for Excel formulas
            if new_value.startswith('='):
                errors.append({
                    'sheet': 'Expression Refs', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Cell contains a formula instead of plain text: "{new_value[:80]}". '
                               f'Formulas cannot be used - please replace with the actual value.',
                })

    # --- Alarm Types sheet ---
    if "Alarm Types" in wb.sheetnames:
        ws = wb["Alarm Types"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 3:
                continue
            name = row[0]
            if not name:
                continue
            new_desc = row[4] if len(row) > 4 else None
            new_msg = row[8] if len(row) > 8 else None
            # Check for formulas in new value cells
            for col_label, val in [('New Description (E)', new_desc), ('New Message (I)', new_msg)]:
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.startswith('='):
                        errors.append({
                            'sheet': 'Alarm Types', 'row': row_idx,
                            'column': col_label,
                            'message': f'Alarm "{name}" cell contains a formula: "{val_str[:80]}". '
                                       f'Formulas cannot be used - please replace with the actual value.',
                        })

    # --- Alarm Priorities sheet ---
    if "Alarm Priorities" in wb.sheetnames:
        ws = wb["Alarm Priorities"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            old_val = str(row[0]).strip() if row[0] else ''
            new_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if old_val and new_val and new_val != old_val:
                # Check for formula in new value
                if new_val.startswith('='):
                    errors.append({
                        'sheet': 'Alarm Priorities', 'row': row_idx,
                        'column': 'New Priority Name (B)',
                        'message': f'Cell contains a formula: "{new_val[:80]}". '
                                   f'Formulas cannot be used - please replace with the actual value.',
                    })

    wb.close()

    if errors:
        log(f"\n{'=' * 60}")
        log(f"Excel Validation: {len(errors)} issue(s) found")
        log(f"{'=' * 60}")
        for e in errors:
            log(f"  [{e['sheet']}] Row {e['row']}, {e['column']}: {e['message']}")
        log(f"{'=' * 60}")
        return False, errors
    else:
        log("Excel validation passed - no issues found.")
        return True, []


def read_lib_edited_excel(excel_path):
    """Read edited FHX Excel. Returns (nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, alarm_changes, priority_changes)."""
    wb = _safe_load_workbook(excel_path)
    nameset_changes = {}
    new_namesets = []
    desc_changes = {}
    sv_changes = []
    expr_changes = []
    alarm_changes = read_alarm_types_excel(wb)
    priority_changes = read_priority_names_excel(wb)

    # Read Namesets sheet
    if "Namesets" in wb.sheetnames:
        ws = wb["Namesets"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            name, category, status, lib_vals_str, new_value = row[:5]
            description = row[5] if len(row) > 5 else ''
            if not name:
                continue

            set_name = str(name).strip()
            status_str = str(status).strip() if status else ''
            category_str = str(category).strip() if category else ''
            desc_str = str(description).strip() if description else ''

            if status_str in ('', 'New'):
                entries_str = new_value if new_value else lib_vals_str
                if not entries_str:
                    continue
                entries_text = [v.strip() for v in str(entries_str).split('\n') if v.strip()]
                entries = []
                for v in entries_text:
                    m = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', v)
                    if m:
                        entries.append({'value': int(m.group(1)), 'name': m.group(2)})
                if entries:
                    new_namesets.append({
                        'name': set_name,
                        'category': category_str,
                        'entries': entries,
                        'description': desc_str,
                    })
            elif new_value:
                lib_vals = [v.strip() for v in str(lib_vals_str).split('\n') if v.strip()] if lib_vals_str else []
                new_vals = [v.strip() for v in str(new_value).split('\n') if v.strip()]
                for i, lib_val in enumerate(lib_vals):
                    if i < len(new_vals) and new_vals[i]:
                        nv = new_vals[i]
                        # Support simple name format (e.g. "STOP") - auto-construct VALUE=... entry
                        if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                            vm = re.match(r'VALUE=(\d+)', lib_val)
                            val_num = vm.group(1) if vm else str(i)
                            nv = f'VALUE={val_num} NAME="{nv}"'
                        nameset_changes[(set_name, lib_val)] = nv

            if desc_str:
                desc_changes[set_name] = desc_str

    # Read String Values sheet
    if "String Values" in wb.sheetnames:
        ws = wb["String Values"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue

            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if new_value and new_value != current_value:
                # Support simple name format (e.g. "STOP")
                nv = new_value
                if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                    vm = re.match(r'VALUE=(\d+)', current_value)
                    val_num = vm.group(1) if vm else '0'
                    nv = f'VALUE={val_num} NAME="{nv}"'
                sv_changes.append((set_name, current_value, nv))

    # Read Expression Refs sheet
    if "Expression Refs" in wb.sheetnames:
        ws = wb["Expression Refs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue

            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if new_value and new_value != current_value:
                # Support simple name format (e.g. "STOP")
                nv = new_value
                if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                    nv = f'NAME="{nv}"'
                expr_changes.append((set_name, current_value, nv))

    wb.close()
    return nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, alarm_changes, priority_changes

def generate_new_lib_fhx(lib_path, setup_path, nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes,
                         output_path, alarm_changes=None, priority_changes=None, log_callback=None, progress_callback=None):
    """Generate new FHX with replaced ENTRY NAME, added ENUMERATION_SET, updated DESCRIPTION, replaced STRING_VALUE, and replaced expression refs."""
    def log(msg):
        if log_callback:
            log_callback(msg)

    def progress(pct, text=''):
        if progress_callback:
            progress_callback(pct, text)

    log(f"Reading FHX: {lib_path}")
    progress(0, 'Reading FHX...')
    lib_content = read_fhx(lib_path)

    log(f"Reading Setup: {setup_path}")
    progress(10, 'Reading Setup...')
    setup_content = read_fhx(setup_path)

    # Extract ENUMERATION_SET definitions for reference
    lib_enum_defs = extract_enum_sets(lib_content)
    setup_enum_defs = extract_enum_sets(setup_content)

    # Build new content
    progress(15, 'Building new FHX...')
    new_content = lib_content

    # Replace LOCALE
    fhx_locale_match = re.search(r'LOCALE="([^"]*)"', new_content)
    setup_locale_match = re.search(r'LOCALE="([^"]*)"', setup_content)
    if fhx_locale_match and setup_locale_match:
        fhx_locale = fhx_locale_match.group(1)
        setup_locale = setup_locale_match.group(1)
        new_content = new_content.replace(f'LOCALE="{fhx_locale}"', f'LOCALE="{setup_locale}"')
        log(f"  LOCALE: {fhx_locale} -> {setup_locale}")

    # Replace ENTRY NAME values in ENUMERATION_SET blocks
    progress(20, 'Replacing ENTRY NAME values...')
    replace_count = 0
    block_cache = {}
    entry_total = len(nameset_changes)
    for idx, ((set_name, old_entry), new_entry) in enumerate(nameset_changes.items()):
        if entry_total > 0 and idx % max(1, entry_total // 20) == 0:
            pct = 20 + int(10 * idx / entry_total)
            progress(pct, f'Replacing ENTRY NAME... ({idx}/{entry_total})')
        old_match = re.search(r'NAME="([^"]*)"', old_entry)
        new_match = re.search(r'NAME="([^"]*)"', new_entry)
        if old_match and new_match:
            old_val_name = old_match.group(1)
            new_val_name = new_match.group(1)
            if old_val_name == new_val_name:
                continue
            if set_name not in block_cache:
                escaped_set = re.escape(set_name)
                block_pattern = f'ENUMERATION_SET\\s+(?:INDEX=\\d+\\s+)?NAME="{escaped_set}"'
                block_match = re.search(block_pattern, new_content)
                if block_match:
                    bstart = block_match.start()
                    brace_pos = new_content.find('{', bstart)
                    if brace_pos >= 0:
                        depth = 1
                        pos = brace_pos + 1
                        while pos < len(new_content) and depth > 0:
                            if new_content[pos] == '{': depth += 1
                            elif new_content[pos] == '}': depth -= 1
                            pos += 1
                        if depth == 0:
                            block_cache[set_name] = (bstart, pos)
            if set_name in block_cache:
                bstart, bpos = block_cache[set_name]
                block = new_content[bstart:bpos]
                escaped_old = re.escape(old_val_name)
                entry_pattern = f'(ENTRY\\s+VALUE=\\d+\\s+NAME="){escaped_old}"'
                new_block, cnt = re.subn(entry_pattern, f'\\g<1>{new_val_name}"', block)
                if cnt > 0:
                    new_content = new_content[:bstart] + new_block + new_content[bpos:]
                    replace_count += cnt
                    log(f"  {set_name}: {old_val_name} -> {new_val_name} ({cnt} occurrences)")
                    block_cache[set_name] = (bstart, bstart + len(new_block))

    # Replace DESCRIPTION in ENUMERATION_SET blocks
    progress(30, 'Replacing descriptions...')
    desc_count = 0
    desc_block_cache = {}
    desc_total = len(desc_changes)
    for idx, (set_name, new_desc) in enumerate(desc_changes.items()):
        if desc_total > 0 and idx % max(1, desc_total // 20) == 0:
            pct = 30 + int(10 * idx / desc_total)
            progress(pct, f'Replacing descriptions... ({idx}/{desc_total})')
        if set_name not in desc_block_cache:
            escaped_set = re.escape(set_name)
            block_pattern = f'ENUMERATION_SET\\s+(?:INDEX=\\d+\\s+)?NAME="{escaped_set}"'
            block_match = re.search(block_pattern, new_content)
            if block_match:
                bstart = block_match.start()
                brace_pos = new_content.find('{', bstart)
                if brace_pos >= 0:
                    depth = 1
                    pos = brace_pos + 1
                    while pos < len(new_content) and depth > 0:
                        if new_content[pos] == '{': depth += 1
                        elif new_content[pos] == '}': depth -= 1
                        pos += 1
                    if depth == 0:
                        desc_block_cache[set_name] = (bstart, pos)
        if set_name in desc_block_cache:
            bstart, bpos = desc_block_cache[set_name]
            block = new_content[bstart:bpos]
            desc_match = re.search(r'DESCRIPTION="[^"]*"', block)
            if desc_match:
                new_block = block[:desc_match.start()] + f'DESCRIPTION="{new_desc}"' + block[desc_match.end():]
            else:
                brace_pos = block.find('{')
                if brace_pos >= 0:
                    insert_pos = brace_pos + 1
                    while insert_pos < len(block) and block[insert_pos] in ' \t\r\n':
                        insert_pos += 1
                    new_block = block[:insert_pos] + f'\r\n  DESCRIPTION="{new_desc}"' + block[insert_pos:]
                else:
                    continue
            new_content = new_content[:bstart] + new_block + new_content[bpos:]
            desc_count += 1
            log(f"  {set_name}: DESCRIPTION=\"{new_desc}\"")
            desc_block_cache[set_name] = (bstart, bstart + len(new_block))

    # Add new namesets
    progress(40, 'Adding new namesets...')
    added_count = 0
    ns_total = len(new_namesets)
    for idx, ns in enumerate(new_namesets):
        if ns_total > 0 and idx % max(1, ns_total // 20) == 0:
            pct = 40 + int(10 * idx / ns_total)
            progress(pct, f'Adding new namesets... ({idx}/{ns_total})')
        name = ns['name']
        category = ns['category']
        entries = ns['entries']
        description = ns.get('description', '')

        block_lines = [f'ENUMERATION_SET NAME="{name}" FIXED=F']
        block_lines.append(f' user="Emerson" time=0')
        block_lines.append('{')
        if description:
            block_lines.append(f'  DESCRIPTION="{description}"')
        if category:
            block_lines.append(f'  CATEGORY="{category}"')
        for entry in entries:
            block_lines.append(f'  ENTRY VALUE={entry["value"]} NAME="{entry["name"]}" {{ }}')
        if entries:
            block_lines.append(f'  DEFAULT_VALUE={entries[0]["value"]}')
        block_lines.append('}')
        block = '\r\n'.join(block_lines)

        insert_idx = new_content.find('FUNCTION_BLOCK_DEFINITIONS')
        if insert_idx < 0:
            insert_idx = new_content.find('FUNCTION_BLOCK ')
        if insert_idx < 0:
            locale_end = new_content.find('\n}\n', new_content.find('LOCALE'))
            if locale_end > 0:
                insert_idx = locale_end + 3
            else:
                insert_idx = len(new_content)
        new_content = new_content[:insert_idx] + block + '\r\n' + new_content[insert_idx:]
        added_count += 1
        log(f"  Added new nameset: {name}")

    # Replace STRING_VALUE references
    progress(50, 'Replacing STRING_VALUE references...')
    sv_count = 0
    sv_total = len(sv_changes)
    for idx, (set_name, old_value, new_value) in enumerate(sv_changes):
        if sv_total > 0 and idx % max(1, sv_total // 20) == 0:
            pct = 50 + int(20 * idx / sv_total)
            progress(pct, f'Replacing STRING_VALUE... ({idx}/{sv_total})')
        new_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_value)
        if new_name_m:
            new_val_name = new_name_m.group(1)
        else:
            new_val_name = new_value

        old_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_value)
        if old_name_m:
            old_val_name = old_name_m.group(1)
        else:
            old_val_name = old_value

        if old_val_name == new_val_name:
            continue

        escaped_set = re.escape(set_name)
        escaped_old = re.escape(old_val_name)
        pattern = f'(SET="{escaped_set}"[^}}]*?STRING_VALUE="){escaped_old}"'
        new_content, count = re.subn(pattern, f'\\g<1>{new_val_name}"', new_content, flags=re.DOTALL)
        if count > 0:
            sv_count += count
            log(f"  {set_name}: {old_val_name} -> {new_val_name} ({count} STRING_VALUE occurrences)")

    # Replace expression references ($nameset:value)
    progress(70, 'Replacing expression references...')
    expr_count = 0
    expr_total = len(expr_changes)
    for idx, (set_name, old_value, new_value) in enumerate(expr_changes):
        if expr_total > 0 and idx % max(1, expr_total // 20) == 0:
            pct = 70 + int(20 * idx / expr_total)
            progress(pct, f'Replacing expression refs... ({idx}/{expr_total})')
        new_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_value)
        if new_name_m:
            new_val_name = new_name_m.group(1)
        else:
            # Also accept NAME="value" format (without VALUE=)
            nm = re.match(r'NAME="([^"]*)"', new_value)
            new_val_name = nm.group(1) if nm else new_value

        old_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_value)
        if old_name_m:
            old_val_name = old_name_m.group(1)
        else:
            old_val_name = old_value

        # Find actual old name in Library FHX ENUMERATION_SET (may be Chinese/mojibake)
        actual_old_name = old_val_name
        if set_name in lib_enum_defs:
            lib_entries = lib_enum_defs[set_name]['entries']
            for entry in lib_entries:
                em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
                if em and em.group(2) == old_val_name:
                    break
                # Match by VALUE number from new_value
                new_vm = re.match(r'VALUE=(\d+)', new_value)
                if new_vm and em and em.group(1) == new_vm.group(1):
                    actual_old_name = em.group(2)
                    break

        if actual_old_name == new_val_name:
            continue

        escaped_set = re.escape(set_name)
        escaped_old = re.escape(actual_old_name)
        # Match '$set_name:old_value' pattern
        pattern = f"'{escaped_set}:{escaped_old}'"
        replacement = f"'{set_name}:{new_val_name}'"
        new_content, count = re.subn(pattern, replacement, new_content)
        if count > 0:
            expr_count += count
            log(f"  {set_name}: {actual_old_name} -> {new_val_name} ({count} expression occurrences)")

    # Additional pass: replace expression values from nameset_changes (ENUMERATION_SET entry translations)
    progress(85, 'Replacing expression refs from nameset changes...')
    if nameset_changes:
        # Build mapping: (set_name, old_value_name) -> new_value_name
        ns_expr_map = {}
        for (ns_name, old_entry), new_entry in nameset_changes.items():
            old_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_entry)
            new_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_entry)
            if old_m and new_m and old_m.group(1) != new_m.group(1):
                ns_expr_map[(ns_name, old_m.group(1))] = new_m.group(1)

        for (ns_name, old_val_name), new_val_name in ns_expr_map.items():
            escaped_set = re.escape(ns_name)
            escaped_old = re.escape(old_val_name)
            pattern = f"'{escaped_set}:{escaped_old}'"
            replacement = f"'{ns_name}:{new_val_name}'"
            new_content, count = re.subn(pattern, replacement, new_content)
            if count > 0:
                expr_count += count
                log(f"  {ns_name}: {old_val_name} -> {new_val_name} ({count} expression refs from nameset change)")

    # Replace alarm values
    alarm_count = 0
    if alarm_changes:
        new_content, alarm_count = replace_alarm_values(new_content, alarm_changes)
        log(f"  Replaced {alarm_count} alarm fields")

    # Replace PRIORITY_NAME values
    priority_count = 0
    if priority_changes:
        new_content, priority_count = replace_priority_names(new_content, priority_changes)
        log(f"  Replaced {priority_count} PRIORITY_NAME values")

    # Write output
    progress(90, 'Writing output...')
    log(f"\nWriting output: {output_path}")
    write_fhx(output_path, new_content)

    progress(100, 'Done')
    log(f"  Done! Replaced {replace_count} ENTRY NAME, Updated {desc_count} descriptions, Added {added_count} new namesets, Replaced {sv_count} STRING_VALUE, Replaced {expr_count} expression refs, Replaced {alarm_count} alarm fields, Translated {priority_count} PRIORITY_NAME")

    return replace_count + desc_count + added_count + sv_count + expr_count + alarm_count + priority_count

# ============================================================
# GUI
# ============================================================
class FHX_Migrator_App:
    def __init__(self, root):
        self.root = root
        self.root.title("DeltaV FHX Nameset Editor")
        self.root.geometry("900x750")
        self.root.resizable(True, True)
        # Set window icon
        try:
            if hasattr(sys, '_MEIPASS'):
                icon_path = os.path.join(sys._MEIPASS, 'exp_logo.ico')
            else:
                icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exp_logo.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass
        self.fhx_path = tk.StringVar()
        self.setup_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self._build_ui()

    def _build_ui(self):
        author_label = tk.Label(self.root, text="Author: Jared.Ji (Jared.Ji@emerson.com)", fg="gray", font=("Arial", 9))
        author_label.pack(side=tk.BOTTOM, pady=(0, 5))

        tab = tk.Frame(self.root)
        tab.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._build_tab(tab)

    def _build_tab(self, parent):
        frame = tk.Frame(parent, padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # Compare section
        cmp_frame = tk.LabelFrame(frame, text="Step 1: Compare FHX with Setup", padx=5, pady=5)
        cmp_frame.pack(fill=tk.X, pady=(0, 5))

        r1 = tk.Frame(cmp_frame)
        r1.pack(fill=tk.X, pady=2)
        tk.Label(r1, text="FHX File:", width=18, anchor='e').pack(side=tk.LEFT)
        self.lib_path = tk.StringVar()
        tk.Entry(r1, textvariable=self.lib_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Button(r1, text="Browse...", command=lambda: self._browse(self.lib_path, "FHX", "*.fhx")).pack(side=tk.LEFT)

        r2 = tk.Frame(cmp_frame)
        r2.pack(fill=tk.X, pady=2)
        tk.Label(r2, text="New Database.fhx:", width=18, anchor='e').pack(side=tk.LEFT)
        self.lib_setup_path = tk.StringVar()
        tk.Entry(r2, textvariable=self.lib_setup_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Button(r2, text="Browse...", command=lambda: self._browse(self.lib_setup_path, "FHX", "*.fhx")).pack(side=tk.LEFT)

        btn1 = tk.Frame(cmp_frame)
        btn1.pack(fill=tk.X, pady=5)
        self.lib_compare_btn = tk.Button(btn1, text="Compare and Export Excel", command=self._do_lib_compare,
                                         bg="#4472C4", fg="white", font=("Arial", 11, "bold"), padx=20, pady=5)
        self.lib_compare_btn.pack(side=tk.LEFT, padx=5)

        # Generate section
        gen_frame = tk.LabelFrame(frame, text="Step 2: Generate New FHX from Edited Excel", padx=5, pady=5)
        gen_frame.pack(fill=tk.X, pady=(0, 5))

        r3 = tk.Frame(gen_frame)
        r3.pack(fill=tk.X, pady=2)
        tk.Label(r3, text="FHX File:", width=18, anchor='e').pack(side=tk.LEFT)
        self.lib_gen_path = tk.StringVar()
        tk.Entry(r3, textvariable=self.lib_gen_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Button(r3, text="Browse...", command=lambda: self._browse(self.lib_gen_path, "FHX", "*.fhx")).pack(side=tk.LEFT)

        r4 = tk.Frame(gen_frame)
        r4.pack(fill=tk.X, pady=2)
        tk.Label(r4, text="New Database.fhx:", width=18, anchor='e').pack(side=tk.LEFT)
        self.lib_gen_setup_path = tk.StringVar()
        tk.Entry(r4, textvariable=self.lib_gen_setup_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Button(r4, text="Browse...", command=lambda: self._browse(self.lib_gen_setup_path, "FHX", "*.fhx")).pack(side=tk.LEFT)

        r5 = tk.Frame(gen_frame)
        r5.pack(fill=tk.X, pady=2)
        tk.Label(r5, text="Edited Excel:", width=18, anchor='e').pack(side=tk.LEFT)
        self.lib_excel_path = tk.StringVar()
        tk.Entry(r5, textvariable=self.lib_excel_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Button(r5, text="Browse...", command=lambda: self._browse(self.lib_excel_path, "Excel", "*.xlsx")).pack(side=tk.LEFT)

        btn2 = tk.Frame(gen_frame)
        btn2.pack(fill=tk.X, pady=5)
        self.lib_generate_btn = tk.Button(btn2, text="Generate New FHX", command=self._do_lib_generate,
                                          bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), padx=20, pady=5)
        self.lib_generate_btn.pack(side=tk.LEFT, padx=5)
        tk.Button(btn2, text="Clear Log", command=lambda: self._clear_log(self.log4)).pack(side=tk.LEFT, padx=5)

        prog_frame = tk.Frame(frame)
        prog_frame.pack(fill=tk.X, pady=2)
        self.prog_label4 = tk.Label(prog_frame, text="", anchor='w')
        self.prog_label4.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress4 = ttk.Progressbar(prog_frame, mode='determinate', length=300)
        self.progress4.pack(side=tk.RIGHT)

        log_frame = tk.LabelFrame(frame, text="Log", padx=5, pady=5)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log4 = scrolledtext.ScrolledText(log_frame, font=("Consolas", 9), wrap=tk.WORD)
        self.log4.pack(fill=tk.BOTH, expand=True)

    def _browse(self, var, file_type, ext):
        path = filedialog.askopenfilename(filetypes=[(f"{file_type} files", ext), ("All files", "*.*")])
        if path:
            var.set(path)

    def _log(self, log_widget, msg):
        log_widget.insert(tk.END, msg + "\n")
        log_widget.see(tk.END)

    def _clear_log(self, log_widget):
        log_widget.delete(1.0, tk.END)

    def _update_progress(self, bar, label, pct, text):
        bar['value'] = pct
        label.config(text=f"{pct}% {text}")

    def _start_bg_task(self, func, *args):
        t = threading.Thread(target=func, args=args, daemon=True)
        t.start()

    def _do_lib_compare(self):
        lib_path = self.lib_path.get().strip()
        setup_path = self.lib_setup_path.get().strip()
        if not lib_path or not setup_path:
            messagebox.showerror("Error", "Please select both FHX and Setup files.")
            return
        if not os.path.exists(lib_path):
            messagebox.showerror("Error", f"FHX not found: {lib_path}")
            return
        if not os.path.exists(setup_path):
            messagebox.showerror("Error", f"Setup not found: {setup_path}")
            return

        base, _ = os.path.splitext(lib_path)
        excel_out = f"{base}_library_comparison.xlsx"
        if os.path.exists(excel_out):
            if not messagebox.askyesno("Confirm", f"Excel file already exists:\n{excel_out}\n\nOverwrite?"):
                return

        self._clear_log(self.log4)
        self._log(self.log4, "DeltaV FHX Nameset Editor - Compare FHX with Setup")
        self._log(self.log4, f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._log(self.log4, "=" * 50)

        self.lib_compare_btn.config(state=tk.DISABLED)
        self._start_bg_task(self._lib_compare_worker, lib_path, setup_path, excel_out)

    def _lib_compare_worker(self, lib_path, setup_path, excel_out):
        try:
            nameset_comp, sv_comp, expr_comp = compare_lib_and_export(
                lib_path, setup_path, excel_out,
                log_callback=lambda m: self.root.after(0, self._log, self.log4, m),
                progress_callback=lambda p, t: self.root.after(0, self._update_progress, self.progress4, self.prog_label4, p, t)
            )

            self.root.after(0, self._log, self.log4, f"\nExcel exported: {excel_out}")
            self.root.after(0, self._log, self.log4, f"  ENUMERATION_SET: {len(nameset_comp)}")
            self.root.after(0, self._log, self.log4, f"  STRING_VALUE: {len(sv_comp)}")
            self.root.after(0, self._log, self.log4, f"  Expression refs: {len(expr_comp)}")
            self.root.after(0, self._log, self.log4, "\nInstructions:")
            self.root.after(0, self._log, self.log4, "  1. Open the Excel file")
            self.root.after(0, self._log, self.log4, "  2. Review 'Namesets' sheet for ENUMERATION_SET definitions")
            self.root.after(0, self._log, self.log4, "  3. Review 'String Values' sheet for STRING_VALUE references")
            self.root.after(0, self._log, self.log4, "  4. Review 'Expression Refs' sheet for expression references")
            self.root.after(0, self._log, self.log4, "  5. Modify 'New Value' columns if needed")
            self.root.after(0, self._log, self.log4, "  6. Go to Step 2 to generate new FHX")

            msg = f"Comparison complete!\n\nENUMERATION_SET: {len(nameset_comp)}\nSTRING_VALUE: {len(sv_comp)}\nExpression refs: {len(expr_comp)}\n\nExcel: {excel_out}"
            self.root.after(0, messagebox.showinfo, "Success", msg)
        except Exception as e:
            self.root.after(0, self._log, self.log4, f"\nERROR: {e}")
            self.root.after(0, messagebox.showerror, "Error", f"Comparison failed:\n{e}")
        finally:
            self.root.after(0, lambda: self.lib_compare_btn.config(state=tk.NORMAL))

    def _do_lib_generate(self):
        lib_path = self.lib_gen_path.get().strip()
        setup_path = self.lib_gen_setup_path.get().strip()
        excel_path = self.lib_excel_path.get().strip()
        if not lib_path or not setup_path or not excel_path:
            messagebox.showerror("Error", "Please select FHX, Setup, and Excel files.")
            return
        if not os.path.exists(lib_path):
            messagebox.showerror("Error", f"FHX not found: {lib_path}")
            return
        if not os.path.exists(setup_path):
            messagebox.showerror("Error", f"Setup not found: {setup_path}")
            return
        if not os.path.exists(excel_path):
            messagebox.showerror("Error", f"Excel not found: {excel_path}")
            return

        base, ext = os.path.splitext(lib_path)
        output_path = f"{base}_NEW{ext}"
        if os.path.exists(output_path):
            if not messagebox.askyesno("Confirm", f"Output exists:\n{output_path}\n\nOverwrite?"):
                return

        self._clear_log(self.log4)
        self._log(self.log4, "DeltaV FHX Nameset Editor - Generate New FHX")
        self._log(self.log4, f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._log(self.log4, "=" * 50)

        self.lib_generate_btn.config(state=tk.DISABLED)
        self._start_bg_task(self._lib_generate_worker, lib_path, setup_path, excel_path, output_path)

    def _lib_generate_worker(self, lib_path, setup_path, excel_path, output_path):
        try:
            # Validate Excel data before proceeding
            self.root.after(0, self._log, self.log4, "Validating Excel data...")
            is_valid, validation_errors = validate_excel_for_generation(
                excel_path,
                log_callback=lambda m: self.root.after(0, self._log, self.log4, m),
            )
            if not is_valid:
                error_summary = '\n'.join(
                    f"[{e['sheet']}] Row {e['row']}, {e['column']}: {e['message']}"
                    for e in validation_errors[:20]
                )
                if len(validation_errors) > 20:
                    error_summary += f"\n... and {len(validation_errors) - 20} more issues"
                self.root.after(0, self._log, self.log4,
                    f"\nGeneration aborted: {len(validation_errors)} issue(s) found in Excel.")
                self.root.after(0, messagebox.showwarning, "Excel Validation Failed",
                    f"Found {len(validation_errors)} issue(s) in Excel:\n\n{error_summary}\n\n"
                    "Please fix these issues and try again.")
                return

            nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, alarm_changes, priority_changes = read_lib_edited_excel(excel_path)
            self.root.after(0, self._log, self.log4, f"Loaded {len(nameset_changes)} value changes, {len(new_namesets)} new namesets, {len(desc_changes)} description changes, {len(sv_changes)} STRING_VALUE changes, {len(expr_changes)} expression changes, {len(alarm_changes)} alarm changes, {len(priority_changes)} priority changes")

            if not nameset_changes and not new_namesets and not desc_changes and not sv_changes and not expr_changes and not alarm_changes and not priority_changes:
                self.root.after(0, messagebox.showinfo, "Info", "No changes found in Excel.")
                return

            count = generate_new_lib_fhx(
                lib_path, setup_path, nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, output_path,
                alarm_changes=alarm_changes, priority_changes=priority_changes,
                log_callback=lambda m: self.root.after(0, self._log, self.log4, m),
                progress_callback=lambda p, t: self.root.after(0, self._update_progress, self.progress4, self.prog_label4, p, t)
            )
            self.root.after(0, self._log, self.log4, f"\n{'=' * 50}")
            self.root.after(0, self._log, self.log4, f"Output: {output_path}")
            self.root.after(0, messagebox.showinfo, "Success",
                f"New FHX generated!\n\n"
                f"Changes: {count}\n"
                f"Output: {output_path}")
        except Exception as e:
            self.root.after(0, self._log, self.log4, f"\nERROR: {e}")
            self.root.after(0, messagebox.showerror, "Error", f"Generation failed:\n{e}")
        finally:
            self.root.after(0, lambda: self.lib_generate_btn.config(state=tk.NORMAL))

# ============================================================
# CLI
# ============================================================
def cli_progress(pct, text=''):
    bar_len = 40
    filled = int(bar_len * pct / 100)
    bar = '#' * filled + '-' * (bar_len - filled)
    print(f'\r  [{bar}] {pct:3d}% {text}', end='', flush=True)
    if pct >= 100:
        print()

def main():
    if len(sys.argv) > 1:
        import argparse
        parser = argparse.ArgumentParser(description='DeltaV FHX Nameset Editor')
        sub = parser.add_subparsers(dest='command')

        # Compare: compare any FHX with Setup, export Excel
        p_compare = sub.add_parser('compare', help='Compare any FHX with Setup, export Excel')
        p_compare.add_argument('fhx', help='Input FHX file (any type)')
        p_compare.add_argument('--setup', help='New Database.fhx reference file', required=True)
        p_compare.add_argument('-o', '--output', help='Output Excel path', default=None)

        # Generate: generate new FHX from edited Excel
        p_generate = sub.add_parser('generate', help='Generate new FHX from edited Excel')
        p_generate.add_argument('fhx', help='Original FHX file')
        p_generate.add_argument('--setup', help='New Database.fhx reference file', required=True)
        p_generate.add_argument('--excel', help='Edited Excel file', required=True)
        p_generate.add_argument('-o', '--output', help='Output FHX path', default=None)

        args = parser.parse_args()

        if args.command == 'compare':
            output = args.output or os.path.splitext(args.fhx)[0] + '_comparison.xlsx'
            nameset_comp, sv_comp, expr_comp = compare_lib_and_export(
                args.fhx, args.setup, output,
                log_callback=print, progress_callback=cli_progress
            )
            print(f"\nExcel exported: {output}")
            print(f"  ENUMERATION_SET: {len(nameset_comp)}")
            print(f"  STRING_VALUE: {len(sv_comp)}")
            print(f"  Expression refs: {len(expr_comp)}")

        elif args.command == 'generate':
            output = args.output or os.path.splitext(args.fhx)[0] + '_NEW' + os.path.splitext(args.fhx)[1]
            # Validate Excel data before proceeding
            print("Validating Excel data...")
            is_valid, validation_errors = validate_excel_for_generation(
                args.excel, log_callback=print
            )
            if not is_valid:
                print(f"\nGeneration aborted: {len(validation_errors)} issue(s) found.")
                sys.exit(1)
            nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, alarm_changes, priority_changes = read_lib_edited_excel(args.excel)
            count = generate_new_lib_fhx(
                args.fhx, args.setup, nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, output,
                alarm_changes=alarm_changes, priority_changes=priority_changes,
                log_callback=print, progress_callback=cli_progress
            )
            print(f"\nOutput: {output}")
            print(f"  Changes: {count}")

        else:
            parser.print_help()
    else:
        root = tk.Tk()
        FHX_Migrator_App(root)
        root.mainloop()

if __name__ == '__main__':
    main()