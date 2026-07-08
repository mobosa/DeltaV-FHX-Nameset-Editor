"""
DeltaV FHX Nameset Core — Backend logic (no GUI dependencies).
Parses, compares, and generates DeltaV FHX configuration files.
Translates nameset values between Chinese and English.
"""

import re
import os
import sys
import time
import json
import zipfile
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# Validation
# ============================================================
def validate_xlsx(filepath):
    """Validate that a file is a valid xlsx (ZIP) file. Returns (ok, error_msg)."""
    if not os.path.exists(filepath):
        return False, f"File not found: {filepath}"
    size = os.path.getsize(filepath)
    if size == 0:
        return False, f"File is empty (0 bytes): {filepath}"
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            bad = zf.testzip()
            if bad is not None:
                return False, f"Corrupt entry in xlsx: {bad}"
        return True, ''
    except zipfile.BadZipFile:
        return False, f"File is not a valid xlsx (not a ZIP file, {size} bytes): {filepath}"

# ============================================================
# Common helpers
# ============================================================
def find_block_end(content, brace_pos):
    """Find matching closing brace starting from an opening brace position.
    Returns the position after the closing brace, or -1 if not found."""
    depth = 1
    pos = brace_pos + 1
    while pos < len(content) and depth > 0:
        if content[pos] == '{':
            depth += 1
        elif content[pos] == '}':
            depth -= 1
        pos += 1
    return pos if depth == 0 else -1

def find_enum_set_block(content, set_name):
    """Find an ENUMERATION_SET block by name.
    Returns (block_start, block_end) or None."""
    escaped = re.escape(set_name)
    pattern = f'ENUMERATION_SET\\s+(?:INDEX=\\d+\\s+)?NAME="{escaped}"'
    m = re.search(pattern, content)
    if not m:
        return None
    brace_pos = content.find('{', m.start())
    if brace_pos < 0:
        return None
    end = find_block_end(content, brace_pos)
    if end < 0:
        return None
    return (m.start(), end)

def _safe_load_workbook(excel_path, log_callback=None):
    """Load an Excel file with retry and fallback. Handles Excel file locking and format issues."""
    last_err = None
    for attempt in range(3):
        try:
            return load_workbook(excel_path, read_only=True)
        except Exception as e:
            last_err = e
            if 'not a zip' in str(e).lower() or 'badzip' in str(e).lower():
                if attempt < 2 and log_callback:
                    log_callback(f"Excel file temporarily unavailable, retrying in 1s (attempt {attempt + 1}/3)...")
                time.sleep(1)
                continue
            raise
    try:
        return load_workbook(excel_path, read_only=False)
    except Exception:
        pass
    try:
        return load_workbook(excel_path, read_only=True, data_only=True)
    except Exception:
        pass
    raise RuntimeError(
        f"Cannot read Excel file. Please make sure:\n"
        f"1. The file is closed in Excel\n"
        f"2. The file is saved as .xlsx format (not .xls)\n"
        f"3. The file is not corrupted\n\n"
        f"Original error: {last_err}"
    )

def _is_chinese(value):
    """Check if a string contains CJK characters (Chinese/Japanese/Korean)."""
    import unicodedata
    for ch in value:
        try:
            name = unicodedata.name(ch, '')
            if 'CJK' in name or 'HIRAGANA' in name or 'KATAKANA' in name:
                return True
        except ValueError:
            continue
    return False

# ============================================================
# File I/O
# ============================================================
def read_fhx(filepath):
    """Read FHX file (UTF-16 LE with BOM or UTF-8)."""
    with open(filepath, 'rb') as f:
        raw = f.read()
    if raw[:2] == b'\xff\xfe':
        return raw[2:].decode('utf-16-le')
    if raw[:2] == b'\xfe\xff':
        return raw[2:].decode('utf-16-be')
    try:
        return raw.decode('utf-8')
    except (UnicodeDecodeError, ValueError):
        return raw.decode('utf-16-le', errors='replace')

def write_fhx(filepath, content):
    """Write FHX file as UTF-16 LE with BOM."""
    with open(filepath, 'wb') as f:
        f.write(b'\xff\xfe')
        f.write(content.encode('utf-16-le'))

# ============================================================
# Extraction helpers
# ============================================================
def extract_enum_sets(content):
    """Extract all ENUMERATION_SET definitions. Returns dict: name -> {'entries': [...], 'category': '...', 'description': '...'}."""
    enum_sets = {}
    for m in re.finditer(r'ENUMERATION_SET\s+(?:INDEX=\d+\s+)?NAME="([^"]*)"', content):
        name = m.group(1)
        start = m.start()
        brace_start = content.find('{', start)
        if brace_start < 0:
            continue
        end = find_block_end(content, brace_start)
        if end < 0:
            continue
        block = content[start:end]
        values = []
        for em in re.finditer(r'ENTRY\s+VALUE=(\d+)\s+NAME="([^"]*)"', block):
            entry_str = f'VALUE={em.group(1)} NAME="{em.group(2)}"'
            values.append(entry_str)
        cat_match = re.search(r'CATEGORY="([^"]*)"', block)
        category = cat_match.group(1) if cat_match else ''
        desc_match = re.search(r'DESCRIPTION="([^"]*)"', block)
        description = desc_match.group(1) if desc_match else ''
        enum_sets[name] = {'entries': values, 'category': category, 'description': description}
    return enum_sets

def extract_string_value_refs(content):
    """Extract SET+STRING_VALUE pairs from FHX. Returns dict: set_name -> {value -> count}."""
    refs = {}
    for m in re.finditer(r'SET="([^"]*)"[^}]*?STRING_VALUE="([^"]*)"', content, re.DOTALL):
        set_name = m.group(1)
        val = m.group(2)
        if set_name not in refs:
            refs[set_name] = {}
        refs[set_name][val] = refs[set_name].get(val, 0) + 1
    return refs

def extract_expression_refs(content):
    """Extract nameset references from expressions. Returns dict: set_name -> {value -> count}."""
    refs = {}
    for m in re.finditer(r"'(\$?[a-zA-Z_][a-zA-Z0-9_-]*):([^']+)'", content):
        set_name = m.group(1)
        val = m.group(2).strip()
        if not val:
            continue
        if set_name not in refs:
            refs[set_name] = {}
        refs[set_name][val] = refs[set_name].get(val, 0) + 1
    return refs

def extract_alarms(content):
    """Extract SYSTEM_ALARM and USER_ALARM blocks. Returns dict: alarm_name -> dict of fields."""
    alarms = {}
    for m in re.finditer(r'(SYSTEM_ALARM|USER_ALARM)\s+(?:INDEX=(\d+)\s+)?NAME="([^"]*)"', content):
        alarm_type = m.group(1)
        alarm_index = m.group(2)
        alarm_name = m.group(3)
        start = m.start()
        brace_start = content.find('{', start)
        if brace_start < 0:
            continue
        end = find_block_end(content, brace_start)
        if end < 0:
            continue
        block = content[start:end]
        fields = {}
        for field in ('DESCRIPTION', 'ALARM_WORD', 'MESSAGE', 'CATEGORY', 'SUMMARY_NO',
                      'DEFAULT_PARAM1', 'DEFAULT_PARAM2', 'WAVE_FILE'):
            fm = re.search(rf'{field}="([^"]*)"', block)
            if fm:
                fields[field] = fm.group(1)
        fields['ALARM_TYPE'] = alarm_type
        if alarm_index:
            fields['INDEX'] = alarm_index
        alarms[alarm_name] = fields
    return alarms

def extract_priority_names(content):
    """Extract unique PRIORITY_NAME / ALARM_ANNUNCIATION UI_NAME values. Returns set of values."""
    names = set()
    for m in re.finditer(r'PRIORITY_NAME="([^"]*)"', content):
        val = m.group(1).strip()
        if val:
            names.add(val)
    for m in re.finditer(r'ALARM_ANNUNCIATION\s+.*?UI_NAME="([^"]*)"', content):
        val = m.group(1).strip()
        if val:
            names.add(val)
    return names

# ============================================================
# Comparison
# ============================================================
def compare_alarms(lib_or_cs_alarms, setup_alarms):
    """Compare alarm definitions. Returns list of dicts for Excel."""
    comparison = []
    all_names = sorted(set(list(lib_or_cs_alarms.keys()) + list(setup_alarms.keys())))
    for name in all_names:
        old_alarm = lib_or_cs_alarms.get(name, {})
        new_alarm = setup_alarms.get(name, {})
        if old_alarm and new_alarm:
            status = 'Both'
        elif old_alarm:
            status = 'Old only'
        else:
            status = 'New only'
        comparison.append({
            'name': name,
            'alarm_type': old_alarm.get('ALARM_TYPE', new_alarm.get('ALARM_TYPE', '')),
            'status': status,
            'old_description': old_alarm.get('DESCRIPTION', ''),
            'new_description': new_alarm.get('DESCRIPTION', ''),
            'old_alarm_word': old_alarm.get('ALARM_WORD', ''),
            'new_alarm_word': new_alarm.get('ALARM_WORD', ''),
            'old_message': old_alarm.get('MESSAGE', ''),
            'new_message': new_alarm.get('MESSAGE', ''),
            'old_category': old_alarm.get('CATEGORY', ''),
            'new_category': new_alarm.get('CATEGORY', ''),
        })
    return comparison

def compare_priority_names(lib_or_cs_names, setup_names, to_chinese=True):
    """Compare PRIORITY_NAME values. Returns list of dicts for Excel."""
    EN_TO_CN = {
        'CRITICAL': chr(0x5371) + chr(0x6025),   # 危急
        'WARNING': chr(0x8B66) + chr(0x544A),    # 警告
        'ADVISORY': chr(0x63D0) + chr(0x793A),   # 提示
        'LOG': chr(0x8BB0) + chr(0x5F55),        # 记录
        'ALERT': chr(0x62A5) + chr(0x8B66),      # 报警
        'INTERLOCK': chr(0x8054) + chr(0x9501),  # 联锁
        'PROMPT': chr(0x63D0) + chr(0x9192),     # 提醒
    }
    CN_TO_EN = {v: k for k, v in EN_TO_CN.items()}
    trans_map = EN_TO_CN if to_chinese else CN_TO_EN
    comparison = []
    for name in sorted(lib_or_cs_names):
        if name in trans_map:
            suggested = trans_map[name]
        else:
            continue
        if suggested == name:
            continue
        comparison.append({
            'old_value': name,
            'new_value': suggested,
            'count': 0,
        })
    return comparison

# ============================================================
# Translation mappings
# ============================================================
DELTA_VALUE_CN_TO_EN = {
    '$phase_state': {
        '停止中': 'Stopping', '正在退出': 'Aborting',
        '空闲': 'Idle', '正在运行': 'Running',
        '已退出': 'Aborted', '已完成': 'Complete',
        '保留中': 'Holding', '正在启动': 'Starting',
        '正在保持': 'Held', '已保留': 'Held',
        '准备就绪': 'Ready', '已停止': 'Stopped',
        '正在重启': 'Restarting', '未加载': 'Not Loaded',
    },
    '$recipe_state': {
        '已完成': 'Complete', '停止中': 'Stopping',
        '正在退出': 'Aborting', '空闲': 'Idle',
        '正在运行': 'Running', '已退出': 'Aborted',
        '保留中': 'Holding', '正在启动': 'Starting',
        '正在保持': 'Held', '已保留': 'Held',
        '准备就绪': 'Ready', '已停止': 'Stopped',
        '正在重启': 'Restarting', '未加载': 'Not Loaded',
    },
    '$sfc_action_states': {
        '已完成': 'Complete', '激活': 'Active',
        '不活动': 'Inactive', '延迟': 'Delayed',
        '待定': 'Pending', '失败': 'Failed',
    },
    '$phase_command': {
        '清除故障': 'Clear Failures', '中止': 'Abort',
        '保持': 'Hold', '停止': 'Stop', '复位': 'Reset',
        '暂停': 'Pause', '下载': 'Download', '恢复': 'Resume',
        '重启': 'Restart', '启动': 'Start',
        '自动': 'Automatic', '手动': 'Manual',
    },
    '$phase_owner_id': {
        'DeltaV 批量': 'DeltaV Batch', '外部': 'External',
    },
    '$sfc_commands': {
        '启动顺控': 'Start Sequence',
        '停止顺控': 'Stop Sequence',
        '复位顺控': 'Reset Sequence',
    },
    '$sfc_states': {
        '顺控空闲': 'Sequence Idle',
        '顺控激活': 'Sequence Active',
        '顺控停止': 'Sequence Stopped',
        '顺控完成': 'Sequence Completed',
        '顺控堵塞': 'Sequence Blocked',
    },
    '$phase_restart_types': {
        '继续': 'Continue', '从启动开始': 'From Start',
        '从启动开始无下载': 'From Start Without Download',
    },
    '$phase_wdog_states': {
        '已失败': 'FAILED', '可疑': 'SUSPECT', '正常': 'OK',
    },
    '$dc_states': {
        '锁定': 'Locked', '已跳车': 'Tripped',
        '非励磁回讯失败': 'Failed Passive',
        '停车/联锁': 'Shutdown/Interlocked',
    },
    '$sysstat_opts': {
        '切换': 'Switchover', '我的下装': 'MyDownload',
        '全部下装': 'TotalDownload', '电源故障': 'Powerfail',
    },
    '$time_format': {
        '本地': 'Local', '协调世界时': 'UTC',
    },
    '$module_states': {
        '离线': 'Out Of Service', '服务中': 'In Service',
    },
    'Alarm_Group_Mode': {
        '已跳车': 'Tripped', '正常': 'Normal', '已停用': 'Disabled',
    },
    'Suppress_Timeout_Opt': {
        '所有已抑制': 'All Suppressed',
        '激活已抑制': 'Active Suppressed',
        '无抑制': 'None Suppressed',
    },
    'LOOP_TYPE': {
        '自动': 'Auto', '手动': 'Manual',
        '结束': 'End', '继续': 'Continue',
    },
    'MAN_RETURN': {
        '自动': 'Auto', '自动回手动': 'Auto To Manual',
        '继续': 'Continue', '重新输入': 'Re-enter',
    },
    'RESTART_MODE': {
        '启动': 'Start', '无等待启动': 'No Wait Start',
        '开始': 'Start', '继续等待开始': 'Continue Wait Start',
    },
    '$isel_typ': {
        '第一个有效值': 'First Valid', '最小': 'Minimum',
        '最大': 'Maximum', '平均': 'Average',
    },
    '$l_typ': {
        '间接': 'Indirect', '直接': 'Direct',
    },
    '$sysstat_ls_opts': {
        '电源故障': 'Powerfail',
    },
    '$time': {
        '秒': 'Seconds', '分': 'Minutes', '时': 'Hours',
    },
}

DELTA_VALUE_EN_TO_CN = {}
for _set, _map in DELTA_VALUE_CN_TO_EN.items():
    for _cn, _en in _map.items():
        DELTA_VALUE_EN_TO_CN.setdefault(_set, {})[_en] = _cn
del _set, _map, _cn, _en


def _bidirectional_translate(set_name, value):
    """Translate a nameset value bidirectionally based on content.
    Chinese value -> English; English value -> Chinese."""
    if _is_chinese(value):
        if set_name in DELTA_VALUE_CN_TO_EN and value in DELTA_VALUE_CN_TO_EN[set_name]:
            return DELTA_VALUE_CN_TO_EN[set_name][value], 'en'
    else:
        if set_name in DELTA_VALUE_EN_TO_CN and value in DELTA_VALUE_EN_TO_CN[set_name]:
            return DELTA_VALUE_EN_TO_CN[set_name][value], 'cn'
    return None, None

# ============================================================
# Excel helpers — generic sheet writer
# ============================================================
_XL_STYLES = {
    'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    'header_font': Font(color='FFFFFF', bold=True, size=11),
    'thin_border': Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    ),
    'yellow_fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    'green_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'red_fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'light_blue_fill': PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid'),
}

def _write_excel_sheet(ws, headers, rows, col_widths=None, data_cols=None, format_func=None):
    """Generic Excel sheet writer with headers and rows.

    Args:
        ws: openpyxl worksheet
        headers: list of header strings
        rows: list of row-data dicts or tuples
        col_widths: dict of column_letter -> width (optional)
        data_cols: list of column keys to extract from each row (optional, for simple 2-col)
        format_func: callable(row_data) -> list of per-cell dicts with 'fill', 'align', 'value' (optional)
    """
    b = _XL_STYLES
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = b['header_fill']
        cell.font = b['header_font']
        cell.alignment = Alignment(horizontal='center')
        cell.border = b['thin_border']

    for i, row_data in enumerate(rows, 2):
        if data_cols and format_func is None:
            # Simple 2-column: data_cols[0] is old, data_cols[1] is new
            for col_idx, key in enumerate(data_cols, 1):
                cell = ws.cell(row=i, column=col_idx, value=row_data.get(key, ''))
                cell.border = b['thin_border']
                if col_idx == 2:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        elif format_func:
            for col_idx, cell_info in enumerate(format_func(row_data), 1):
                cell = ws.cell(row=i, column=col_idx, value=cell_info.get('value', ''))
                cell.border = b['thin_border']
                if 'fill' in cell_info and cell_info['fill']:
                    cell.fill = cell_info['fill']
                wrap = cell_info.get('wrap', col_idx > 1)
                cell.alignment = Alignment(wrap_text=wrap, vertical='top')

    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'
    if rows:
        last_col = chr(ord('A') + len(headers) - 1)
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

# ============================================================
# Excel read/write for each sheet type
# ============================================================
def write_alarm_types_sheet(wb, alarm_comparison):
    """Write the 'Alarm Types' sheet to an existing workbook."""
    if not alarm_comparison:
        return
    ws = wb.create_sheet("Alarm Types")
    b = _XL_STYLES
    headers = ['Alarm Name', 'Alarm Type', 'Status',
               'Old Description', 'New Description',
               'Old Alarm Word', 'New Alarm Word',
               'Old Message', 'New Message',
               'Old Category', 'New Category']

    def _fmt(item):
        status = item['status']
        status_fill = b['green_fill'] if status == 'Both' else (
            b['yellow_fill'] if status == 'Old only' else b['red_fill'])
        return [
            {'value': item['name']},
            {'value': item.get('alarm_type', '')},
            {'value': status, 'fill': status_fill},
            {'value': item.get('old_description', '')},
            {'value': item.get('new_description', '')},
            {'value': item.get('old_alarm_word', '')},
            {'value': item.get('new_alarm_word', '')},
            {'value': item.get('old_message', '')},
            {'value': item.get('new_message', '')},
            {'value': item.get('old_category', '')},
            {'value': item.get('new_category', '')},
        ]

    _write_excel_sheet(ws, headers, alarm_comparison, format_func=_fmt,
                       col_widths={'A': 30, 'B': 15, 'C': 12, 'D': 30, 'E': 30,
                                   'F': 20, 'G': 20, 'H': 40, 'I': 40, 'J': 15, 'K': 15})

def read_alarm_types_excel(wb):
    """Read 'Alarm Types' sheet from workbook. Returns dict: alarm_name -> dict of new field values."""
    alarm_changes = {}
    if "Alarm Types" not in wb.sheetnames:
        return alarm_changes
    ws = wb["Alarm Types"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        new_description = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        new_alarm_word = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        new_message = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        new_category = str(row[10]).strip() if len(row) > 10 and row[10] else ''
        changes = {}
        if new_description:
            changes['DESCRIPTION'] = new_description
        if new_alarm_word:
            changes['ALARM_WORD'] = new_alarm_word
        if new_message:
            changes['MESSAGE'] = new_message
        if new_category:
            changes['CATEGORY'] = new_category
        if changes:
            alarm_changes[name] = changes
    return alarm_changes

def write_priority_names_sheet(wb, priority_comparison):
    """Write 'Alarm Priorities' sheet to workbook."""
    if not priority_comparison:
        return
    ws = wb.create_sheet("Alarm Priorities")
    b = _XL_STYLES
    headers = ['Old Priority Name', 'New Priority Name']

    def _fmt(item):
        new_fill = b['yellow_fill'] if not item['new_value'] else None
        return [
            {'value': item['old_value']},
            {'value': item['new_value'], 'fill': new_fill},
        ]

    _write_excel_sheet(ws, headers, priority_comparison, format_func=_fmt,
                       col_widths={'A': 30, 'B': 30})

def read_priority_names_excel(wb):
    """Read 'Alarm Priorities' sheet. Returns dict: old_value -> new_value."""
    changes = {}
    if "Alarm Priorities" not in wb.sheetnames:
        return changes
    ws = wb["Alarm Priorities"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        old_val = str(row[0]).strip() if row[0] else ''
        new_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if old_val and new_val and old_val != new_val:
            changes[old_val] = new_val
    return changes

def replace_priority_names(content, priority_changes):
    """Replace PRIORITY_NAME values and ALARM_ANNUNCIATION UI_NAME values."""
    count = 0
    for old_val, new_val in priority_changes.items():
        old_pattern = f'PRIORITY_NAME="{old_val}"'
        new_pattern = f'PRIORITY_NAME="{new_val}"'
        n = content.count(old_pattern)
        content = content.replace(old_pattern, new_pattern)
        count += n
        old_ui = f'UI_NAME="{old_val}"'
        new_ui = f'UI_NAME="{new_val}"'
        n2 = content.count(old_ui)
        content = content.replace(old_ui, new_ui)
        count += n2
    return content, count

def replace_alarm_values(content, alarm_changes):
    """Replace DESCRIPTION, ALARM_WORD, MESSAGE, CATEGORY in SYSTEM_ALARM/USER_ALARM blocks."""
    replace_count = 0
    for alarm_name, changes in alarm_changes.items():
        escaped_name = re.escape(alarm_name)
        pattern = r'(SYSTEM_ALARM|USER_ALARM)\s+(?:INDEX=\d+\s+)?NAME="' + escaped_name + r'"'
        m = re.search(pattern, content)
        if not m:
            continue
        start = m.start()
        brace_pos = content.find('{', start)
        if brace_pos < 0:
            continue
        end = find_block_end(content, brace_pos)
        if end < 0:
            continue
        block = content[start:end]
        for field, new_val in changes.items():
            fm = re.search(rf'{field}="[^"]*"', block)
            if fm:
                old_val = fm.group(0)
                new_field = f'{field}="{new_val}"'
                block = block.replace(old_val, new_field, 1)
                replace_count += 1
        content = content[:start] + block + content[end:]
    return content, replace_count

# ============================================================
# Compare & Export (Step 1)
# ============================================================
def compare_lib_and_export(lib_path, setup_path, output_path, log_callback=None, progress_callback=None):
    """Compare FHX ENUMERATION_SET + STRING_VALUE + Expression refs with Setup. Export Excel."""
    def log(msg):
        if log_callback:
            log_callback(msg)
    def progress(pct, text=''):
        if progress_callback:
            progress_callback(pct, text)

    log(f"Reading FHX: {lib_path}")
    progress(0, 'Reading FHX...')
    lib_content = read_fhx(lib_path)
    log(f"  {len(lib_content):,} chars")

    log(f"Reading Setup: {setup_path}")
    progress(15, 'Reading Setup...')
    setup_content = read_fhx(setup_path)
    log(f"  {len(setup_content):,} chars")

    progress(25, 'Extracting ENUMERATION_SET definitions...')
    lib_enum_defs = extract_enum_sets(lib_content)
    log(f"  FHX ENUMERATION_SET: {len(lib_enum_defs)}")
    setup_enum_defs = extract_enum_sets(setup_content)
    log(f"  Setup ENUMERATION_SET: {len(setup_enum_defs)}")

    progress(35, 'Extracting STRING_VALUE references...')
    lib_refs = extract_string_value_refs(lib_content)
    log(f"  FHX SET+STRING_VALUE refs: {len(lib_refs)}")

    progress(45, 'Extracting expression references...')
    expr_refs = extract_expression_refs(lib_content)
    log(f"  FHX expression refs: {len(expr_refs)}")

    # Build ENUMERATION_SET comparison
    progress(50, 'Building ENUMERATION_SET comparison...')
    nameset_comparison = []
    for set_name in sorted(lib_enum_defs.keys()):
        lib_data = lib_enum_defs[set_name]
        lib_vals = lib_data['entries'] if isinstance(lib_data, dict) else lib_data
        lib_cat = lib_data['category'] if isinstance(lib_data, dict) else ''
        lib_desc = lib_data['description'] if isinstance(lib_data, dict) else ''

        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'category': '', 'description': ''})
        setup_vals = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_cat = setup_data['category'] if isinstance(setup_data, dict) else ''

        if set_name in setup_enum_defs:
            status = 'Both'
            default_suggestion = '\n'.join(setup_vals) if setup_vals else ''
            category = setup_cat
            value_count_mismatch = len(lib_vals) != len(setup_vals)
        else:
            status = 'Original only'
            default_suggestion = ''
            category = lib_cat
            value_count_mismatch = False

        nameset_comparison.append({
            'name': set_name,
            'fhx_values': lib_vals,
            'status': status,
            'default_suggestion': default_suggestion,
            'category': category,
            'description': lib_desc,
            'value_count_mismatch': value_count_mismatch,
        })

    # Build STRING_VALUE comparison
    progress(60, 'Building STRING_VALUE comparison...')
    sv_comparison = []
    for set_name in sorted(lib_refs.keys()):
        vals = lib_refs[set_name]
        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'description': ''})
        setup_entries = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_desc = setup_data['description'] if isinstance(setup_data, dict) else ''

        entry_map = {}
        for entry in setup_entries:
            em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
            if em:
                entry_map[em.group(2)] = entry

        for val in sorted(vals.keys()):
            count = vals[val]
            suggested = ''
            for entry_name, entry_str in entry_map.items():
                if val == entry_name:
                    suggested = entry_str
                    break
            if not suggested:
                val_m = re.match(r'VALUE=(\d+)', val)
                if val_m:
                    for entry_name, entry_str in entry_map.items():
                        if f'VALUE={val_m.group(1)}' in entry_str:
                            suggested = entry_str
                            break
            if not suggested and set_name in DELTA_VALUE_CN_TO_EN:
                trans_val, direction = _bidirectional_translate(set_name, val)
                if trans_val is not None and trans_val in entry_map:
                    suggested = entry_map[trans_val]

            sv_comparison.append({
                'set_name': set_name,
                'current_value': val,
                'count': count,
                'suggested': suggested,
                'description': setup_desc,
                'setup_entries': '\n'.join(setup_entries),
            })

    # Build expression comparison
    progress(70, 'Building expression comparison...')
    expr_comparison = []
    for set_name in sorted(expr_refs.keys()):
        vals = expr_refs[set_name]
        setup_data = setup_enum_defs.get(set_name, {'entries': [], 'description': ''})
        setup_entries = setup_data['entries'] if isinstance(setup_data, dict) else setup_data
        setup_desc = setup_data['description'] if isinstance(setup_data, dict) else ''

        entry_map = {}
        for entry in setup_entries:
            em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
            if em:
                entry_map[em.group(2)] = entry

        lib_entry_map = {}
        lib_name_to_num = {}
        lib_num_to_entry = {}
        if set_name in lib_enum_defs:
            for entry in lib_enum_defs[set_name]['entries']:
                em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
                if em:
                    lib_entry_map[em.group(2)] = entry
                    lib_name_to_num[em.group(2)] = em.group(1)
                    lib_num_to_entry[em.group(1)] = entry

        setup_num_to_name = {}
        for entry_name, entry_str in entry_map.items():
            vm = re.match(r'VALUE=(\d+)', entry_str)
            if vm:
                setup_num_to_name[vm.group(1)] = entry_name

        for val in sorted(vals.keys()):
            count = vals[val]
            suggested = ''
            for entry_name, entry_str in entry_map.items():
                if val == entry_name:
                    suggested = entry_str
                    break
            if not suggested and val in lib_name_to_num:
                val_num = lib_name_to_num[val]
                if val_num in setup_num_to_name:
                    eng_name = setup_num_to_name[val_num]
                    suggested = f'VALUE={val_num} NAME="{eng_name}"'
            if not suggested and val in lib_entry_map:
                suggested = lib_entry_map[val]
            if not suggested and set_name in DELTA_VALUE_CN_TO_EN:
                trans_val, direction = _bidirectional_translate(set_name, val)
                if trans_val is not None:
                    if trans_val in entry_map:
                        suggested = entry_map[trans_val]
                    elif setup_num_to_name:
                        for num, name in setup_num_to_name.items():
                            if name == trans_val:
                                suggested = f'VALUE={num} NAME="{trans_val}"'
                                break

            expr_comparison.append({
                'set_name': set_name,
                'current_value': val,
                'count': count,
                'suggested': suggested,
                'description': setup_desc,
                'setup_entries': '\n'.join(setup_entries) if setup_entries else
                    '\n'.join(lib_enum_defs.get(set_name, {}).get('entries', [])),
            })

    # Extract and compare alarm types
    progress(78, 'Extracting alarm types...')
    lib_alarms = extract_alarms(lib_content)
    setup_alarms = extract_alarms(setup_content)
    alarm_comparison = compare_alarms(lib_alarms, setup_alarms)
    log(f"  FHX alarms: {len(lib_alarms)}, Setup alarms: {len(setup_alarms)}")

    lib_locale_m = re.search(r'LOCALE="([^"]*)"', lib_content)
    lib_locale = lib_locale_m.group(1) if lib_locale_m else ''
    log(f"  LOCALE: {lib_locale}")
    is_english_locale = 'english' in lib_locale.lower()
    is_chinese_locale = 'chinese' in lib_locale.lower()

    if is_english_locale or is_chinese_locale:
        lib_priorities = extract_priority_names(lib_content)
        setup_priorities = extract_priority_names(setup_content)
        priority_comparison = compare_priority_names(
            lib_priorities, setup_priorities, to_chinese=is_english_locale)
        for item in priority_comparison:
            item['count'] = lib_content.count(f'PRIORITY_NAME="{item["old_value"]}"')
        log(f"  FHX priorities: {len(lib_priorities)}, Setup priorities: {len(setup_priorities)}")
    else:
        priority_comparison = []
        log("  Priority names: skipped (unsupported locale)")

    progress(80, 'Writing Excel...')
    write_lib_comparison_excel(nameset_comparison, sv_comparison, expr_comparison,
                               output_path, alarm_comparison, priority_comparison)

    progress(100, 'Done')
    log(f"\nExcel exported: {output_path}")
    log(f"  ENUMERATION_SET definitions: {len(nameset_comparison)}")
    log(f"  Alarm types: {len(alarm_comparison)}")
    log(f"  STRING_VALUE references: {len(sv_comparison)}")
    log(f"  Expression references: {len(expr_comparison)}")

    log("\n" + "=" * 60)
    log("ENUMERATION_SET DEFINITIONS:")
    log("=" * 60)
    for item in nameset_comparison:
        log(f"\n[{item['name']}] ({item['status']})")
        if item['fhx_values']:
            for val in item['fhx_values']:
                log(f"  Lib: {val}")
        if item['default_suggestion']:
            for val in item['default_suggestion'].split('\n'):
                if val.strip():
                    log(f"  Setup: {val.strip()}")

    log("\n" + "=" * 60)
    log("STRING_VALUE REFERENCES:")
    log("=" * 60)
    current_set = ''
    for item in sv_comparison:
        if item['set_name'] != current_set:
            current_set = item['set_name']
            log(f"\n[{current_set}]")
        count_str = f" (x{item['count']})" if item['count'] > 1 else ""
        suggested_str = f" -> {item['suggested']}" if item['suggested'] else " (no match)"
        log(f"  {item['current_value']}{count_str}{suggested_str}")

    log("\n" + "=" * 60)
    log("EXPRESSION REFERENCES:")
    log("=" * 60)
    current_set = ''
    for item in expr_comparison:
        if item['set_name'] != current_set:
            current_set = item['set_name']
            log(f"\n[{current_set}]")
        count_str = f" (x{item['count']})" if item['count'] > 1 else ""
        suggested_str = f" -> {item['suggested']}" if item['suggested'] else " (no match)"
        log(f"  {item['current_value']}{count_str}{suggested_str}")

    return nameset_comparison, sv_comparison, expr_comparison

def write_lib_comparison_excel(nameset_comparison, sv_comparison, expr_comparison,
                                output_path, alarm_comparison=None, priority_comparison=None):
    """Write FHX comparison to Excel with five sheets."""
    wb = Workbook()
    b = _XL_STYLES

    # Sheet 1: ENUMERATION_SET definitions
    ws1 = wb.active
    ws1.title = "Namesets"
    headers1 = ['SET Name', 'Category', 'Status', 'FHX Values', 'New Value', 'Description']

    def _fmt_nameset(item):
        status = item['status']
        status_fill = b['green_fill'] if status == 'Both' else b['red_fill']
        new_fill = b['light_blue_fill'] if item['default_suggestion'] else b['yellow_fill']
        fhx_fill = b['yellow_fill'] if item.get('value_count_mismatch') else None
        return [
            {'value': item['name']},
            {'value': item.get('category', '')},
            {'value': status, 'fill': status_fill},
            {'value': '\n'.join(item['fhx_values']), 'fill': fhx_fill},
            {'value': item['default_suggestion'], 'fill': new_fill},
            {'value': item.get('description', '')},
        ]
    _write_excel_sheet(ws1, headers1, nameset_comparison, format_func=_fmt_nameset,
                       col_widths={'A': 30, 'B': 25, 'C': 15, 'D': 40, 'E': 40, 'F': 40})

    # Sheet 2: STRING_VALUE references
    ws2 = wb.create_sheet("String Values")
    headers2 = ['SET Name', 'Current Value', 'Count', 'New Value', 'Setup Entries', 'Description']

    def _fmt_sv(item):
        new_fill = b['light_blue_fill'] if item['suggested'] else b['yellow_fill']
        return [
            {'value': item['set_name']},
            {'value': item['current_value']},
            {'value': item['count']},
            {'value': item['suggested'], 'fill': new_fill},
            {'value': item['setup_entries']},
            {'value': item.get('description', '')},
        ]
    _write_excel_sheet(ws2, headers2, sv_comparison, format_func=_fmt_sv,
                       col_widths={'A': 30, 'B': 30, 'C': 8, 'D': 40, 'E': 50, 'F': 40})

    # Sheet 3: Expression references
    ws3 = wb.create_sheet("Expression Refs")
    _write_excel_sheet(ws3, headers2, expr_comparison, format_func=_fmt_sv,
                       col_widths={'A': 30, 'B': 30, 'C': 8, 'D': 40, 'E': 50, 'F': 40})

    write_alarm_types_sheet(wb, alarm_comparison)
    write_priority_names_sheet(wb, priority_comparison)

    wb.save(output_path)
    ok, err = validate_xlsx(output_path)
    if not ok:
        raise RuntimeError(f"Failed to write valid Excel file: {err}")

# ============================================================
# Excel Validation
# ============================================================
def validate_excel_for_generation(excel_path, log_callback=None):
    """Validate Excel data before generating new FHX. Returns (is_valid, errors)."""
    def log(msg):
        if log_callback:
            log_callback(msg)

    errors = []
    wb = _safe_load_workbook(excel_path, log_callback=log_callback)

    # --- Namesets sheet ---
    if "Namesets" in wb.sheetnames:
        ws = wb["Namesets"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 5:
                continue
            name, category, status, lib_vals_str, new_value = row[:5]
            if not name:
                continue
            set_name = str(name).strip()
            status_str = str(status).strip() if status else ''

            if new_value:
                nv_str = str(new_value).strip()
                if nv_str.startswith('='):
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'Cell contains a formula instead of plain text: "{nv_str[:80]}". '
                                   f'Formulas cannot be used - please replace with the actual value.',
                    })

            if status_str in ('', 'New'):
                entries_str = new_value if new_value else lib_vals_str
                if not entries_str:
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'New nameset "{set_name}" has no entries in New Value or FHX Values column',
                    })
                    continue
                entries_text = [v.strip() for v in str(entries_str).split('\n') if v.strip()]
                valid_count = 0
                seen_nums = {}
                for line_idx, line in enumerate(entries_text):
                    m = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', line)
                    if m:
                        val_num = int(m.group(1))
                        val_name = m.group(2)
                        if not val_name:
                            errors.append({
                                'sheet': 'Namesets', 'row': row_idx,
                                'column': 'New Value (E)',
                                'message': f'Entry has empty NAME: VALUE={val_num} NAME=""',
                            })
                        if val_num in seen_nums:
                            errors.append({
                                'sheet': 'Namesets', 'row': row_idx,
                                'column': 'New Value (E)',
                                'message': f'Duplicate VALUE number {val_num} in nameset "{set_name}" '
                                           f'(first at line {seen_nums[val_num]}, again at line {line_idx + 1})',
                            })
                        else:
                            seen_nums[val_num] = line_idx + 1
                        valid_count += 1
                    else:
                        errors.append({
                            'sheet': 'Namesets', 'row': row_idx,
                            'column': 'New Value (E)',
                            'message': f'Invalid entry format: "{line}" (expected VALUE=N NAME="...")',
                        })
                if valid_count == 0 and entries_text:
                    errors.append({
                        'sheet': 'Namesets', 'row': row_idx,
                        'column': 'New Value (E)',
                        'message': f'No valid VALUE=N NAME="..." entries found for new nameset "{set_name}"',
                    })
            else:
                if new_value:
                    nv_str = str(new_value).strip()
                    nv_lines = [l.strip() for l in nv_str.split('\n') if l.strip()]
                    lib_lines = [l.strip() for l in str(lib_vals_str).split('\n') if l.strip()] if lib_vals_str else []
                    for line_idx, nv_line in enumerate(nv_lines):
                        if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*")', nv_line):
                            if ' ' in nv_line and not nv_line.startswith('VALUE='):
                                errors.append({
                                    'sheet': 'Namesets', 'row': row_idx,
                                    'column': 'New Value (E)',
                                    'message': f'Dubious entry format (has spaces): "{nv_line}" - '
                                               f'Expected VALUE=N NAME="..." or simple name',
                                })

    # --- String Values sheet ---
    if "String Values" in wb.sheetnames:
        ws = wb["String Values"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if not new_value or new_value == current_value:
                continue
            if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*"|[^\s]+)', new_value):
                errors.append({
                    'sheet': 'String Values', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Invalid format: "{new_value}" for set "{set_name}", value "{current_value}"',
                })
            if new_value.startswith('='):
                errors.append({
                    'sheet': 'String Values', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Cell contains a formula instead of plain text: "{new_value[:80]}". '
                               f'Formulas cannot be used - please replace with the actual value.',
                })

    # --- Expression Refs sheet ---
    if "Expression Refs" in wb.sheetnames:
        ws = wb["Expression Refs"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''

            if not new_value or new_value == current_value:
                continue
            if not re.match(r'(?:VALUE=\d+\s+NAME="[^"]*"|NAME="[^"]*"|[^\s]+)', new_value):
                errors.append({
                    'sheet': 'Expression Refs', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Invalid format: "{new_value}" for set "{set_name}", value "{current_value}"',
                })
            if new_value.startswith('='):
                errors.append({
                    'sheet': 'Expression Refs', 'row': row_idx,
                    'column': 'New Value (D)',
                    'message': f'Cell contains a formula instead of plain text: "{new_value[:80]}". '
                               f'Formulas cannot be used - please replace with the actual value.',
                })

    # --- Alarm Types sheet ---
    if "Alarm Types" in wb.sheetnames:
        ws = wb["Alarm Types"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 3:
                continue
            name = row[0]
            if not name:
                continue
            new_desc = row[4] if len(row) > 4 else None
            new_msg = row[8] if len(row) > 8 else None
            for col_label, val in [('New Description (E)', new_desc), ('New Message (I)', new_msg)]:
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.startswith('='):
                        errors.append({
                            'sheet': 'Alarm Types', 'row': row_idx,
                            'column': col_label,
                            'message': f'Alarm "{name}" cell contains a formula: "{val_str[:80]}". '
                                       f'Formulas cannot be used - please replace with the actual value.',
                        })

    # --- Alarm Priorities sheet ---
    if "Alarm Priorities" in wb.sheetnames:
        ws = wb["Alarm Priorities"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            old_val = str(row[0]).strip() if row[0] else ''
            new_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if old_val and new_val and new_val != old_val:
                if new_val.startswith('='):
                    errors.append({
                        'sheet': 'Alarm Priorities', 'row': row_idx,
                        'column': 'New Priority Name (B)',
                        'message': f'Cell contains a formula: "{new_val[:80]}". '
                                   f'Formulas cannot be used - please replace with the actual value.',
                    })

    wb.close()

    if errors:
        log(f"\n{'=' * 60}")
        log(f"Excel Validation: {len(errors)} issue(s) found")
        log(f"{'=' * 60}")
        for e in errors:
            log(f"  [{e['sheet']}] Row {e['row']}, {e['column']}: {e['message']}")
        log(f"{'=' * 60}")
        return False, errors
    else:
        log("Excel validation passed - no issues found.")
        return True, []

# ============================================================
# Read Edited Excel (Step 2 input)
# ============================================================
def read_lib_edited_excel(excel_path):
    """Read edited FHX Excel. Returns 7-tuple of change data."""
    wb = _safe_load_workbook(excel_path)
    nameset_changes = {}
    new_namesets = []
    desc_changes = {}
    sv_changes = []
    expr_changes = []
    alarm_changes = read_alarm_types_excel(wb)
    priority_changes = read_priority_names_excel(wb)

    if "Namesets" in wb.sheetnames:
        ws = wb["Namesets"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            name, category, status, lib_vals_str, new_value = row[:5]
            description = row[5] if len(row) > 5 else ''
            if not name:
                continue

            set_name = str(name).strip()
            status_str = str(status).strip() if status else ''
            category_str = str(category).strip() if category else ''
            desc_str = str(description).strip() if description else ''

            if status_str in ('', 'New'):
                entries_str = new_value if new_value else lib_vals_str
                if not entries_str:
                    continue
                entries_text = [v.strip() for v in str(entries_str).split('\n') if v.strip()]
                entries = []
                for v in entries_text:
                    m = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', v)
                    if m:
                        entries.append({'value': int(m.group(1)), 'name': m.group(2)})
                if entries:
                    new_namesets.append({
                        'name': set_name,
                        'category': category_str,
                        'entries': entries,
                        'description': desc_str,
                    })
            elif new_value:
                lib_vals = [v.strip() for v in str(lib_vals_str).split('\n') if v.strip()] if lib_vals_str else []
                new_vals = [v.strip() for v in str(new_value).split('\n') if v.strip()]
                for i, lib_val in enumerate(lib_vals):
                    if i < len(new_vals) and new_vals[i]:
                        nv = new_vals[i]
                        if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                            vm = re.match(r'VALUE=(\d+)', lib_val)
                            val_num = vm.group(1) if vm else str(i)
                            nv = f'VALUE={val_num} NAME="{nv}"'
                        nameset_changes[(set_name, lib_val)] = nv

            if desc_str:
                desc_changes[set_name] = desc_str

    if "String Values" in wb.sheetnames:
        ws = wb["String Values"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''
            if new_value and new_value != current_value:
                nv = new_value
                if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                    vm = re.match(r'VALUE=(\d+)', current_value)
                    val_num = vm.group(1) if vm else '0'
                    nv = f'VALUE={val_num} NAME="{nv}"'
                sv_changes.append((set_name, current_value, nv))

    if "Expression Refs" in wb.sheetnames:
        ws = wb["Expression Refs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            set_name, current_value, count, new_value = row[:4]
            if not set_name or not current_value:
                continue
            set_name = str(set_name).strip()
            current_value = str(current_value).strip()
            new_value = str(new_value).strip() if new_value else ''
            if new_value and new_value != current_value:
                nv = new_value
                if not re.match(r'VALUE=\d+\s+NAME="[^"]*"', nv):
                    nv = f'NAME="{nv}"'
                expr_changes.append((set_name, current_value, nv))

    wb.close()
    return nameset_changes, new_namesets, desc_changes, sv_changes, expr_changes, alarm_changes, priority_changes

# ============================================================
# Generate New FHX (Step 2)
# ============================================================
def generate_new_lib_fhx(lib_path, setup_path, nameset_changes, new_namesets, desc_changes,
                         sv_changes, expr_changes, output_path,
                         alarm_changes=None, priority_changes=None,
                         log_callback=None, progress_callback=None):
    """Generate new FHX with replaced values. Returns total change count."""
    def log(msg):
        if log_callback:
            log_callback(msg)
    def progress(pct, text=''):
        if progress_callback:
            progress_callback(pct, text)

    log(f"Reading FHX: {lib_path}")
    progress(0, 'Reading FHX...')
    lib_content = read_fhx(lib_path)
    log(f"  {len(lib_content):,} chars")

    log(f"Reading Setup: {setup_path}")
    progress(10, 'Reading Setup...')
    setup_content = read_fhx(setup_path)
    log(f"  {len(setup_content):,} chars")

    lib_enum_defs = extract_enum_sets(lib_content)
    setup_enum_defs = extract_enum_sets(setup_content)
    log(f"  FHX ENUMERATION_SET: {len(lib_enum_defs)}, Setup: {len(setup_enum_defs)}")

    progress(15, 'Building new FHX...')
    new_content = lib_content

    # Replace LOCALE
    fhx_locale_match = re.search(r'LOCALE="([^"]*)"', new_content)
    setup_locale_match = re.search(r'LOCALE="([^"]*)"', setup_content)
    if fhx_locale_match and setup_locale_match:
        fhx_locale = fhx_locale_match.group(1)
        setup_locale = setup_locale_match.group(1)
        new_content = new_content.replace(f'LOCALE="{fhx_locale}"', f'LOCALE="{setup_locale}"')
        log(f"  LOCALE: {fhx_locale} -> {setup_locale}")

    # Replace ENTRY NAME values in ENUMERATION_SET blocks
    progress(20, 'Replacing ENTRY NAME values...')
    replace_count = 0
    block_mods = {}
    for (set_name, old_entry), new_entry in nameset_changes.items():
        old_match = re.search(r'NAME="([^"]*)"', old_entry)
        new_match = re.search(r'NAME="([^"]*)"', new_entry)
        if old_match and new_match:
            old_val_name = old_match.group(1)
            new_val_name = new_match.group(1)
            if old_val_name != new_val_name:
                block_mods.setdefault(set_name, []).append((old_val_name, new_val_name))

    if block_mods:
        block_positions = {}
        for set_name in block_mods:
            pos = find_enum_set_block(new_content, set_name)
            if pos:
                block_positions[set_name] = pos

        modified_blocks = {}
        block_list = list(block_mods.items())
        for idx, (set_name, mods) in enumerate(block_list):
            if set_name not in block_positions:
                continue
            bstart, bpos = block_positions[set_name]
            block = new_content[bstart:bpos]
            for old_name, new_name in mods:
                escaped = re.escape(old_name)
                pattern = f'(ENTRY\\s+VALUE=\\d+\\s+NAME="){escaped}"'
                block, cnt = re.subn(pattern, f'\\g<1>{new_name}"', block)
                replace_count += cnt
                log(f"  {set_name}: {old_name} -> {new_name} ({cnt} occurrences)")
            modified_blocks[set_name] = (bstart, bpos, block)

        sorted_blocks = sorted(modified_blocks.items(), key=lambda x: x[1][0], reverse=True)
        for idx, (set_name, (bstart, bpos, new_block)) in enumerate(sorted_blocks):
            if idx % 10 == 0:
                progress(20 + int(8 * idx / max(1, len(sorted_blocks))),
                         f'Reassembling FHX... ({idx}/{len(sorted_blocks)})')
            log(f"  [{idx+1}/{len(sorted_blocks)}] ENTRY NAME {set_name}")
            new_content = new_content[:bstart] + new_block + new_content[bpos:]
        progress(28, f'Reassembled {len(sorted_blocks)} blocks')

    # Replace DESCRIPTION in ENUMERATION_SET blocks
    progress(30, 'Replacing descriptions...')
    desc_count = 0
    desc_block_mods = {}
    for set_name, new_desc in desc_changes.items():
        desc_block_mods[set_name] = new_desc

    if desc_block_mods:
        desc_positions = {}
        for set_name in desc_block_mods:
            pos = find_enum_set_block(new_content, set_name)
            if pos:
                desc_positions[set_name] = pos

        modified_desc_blocks = {}
        for set_name, new_desc in desc_block_mods.items():
            if set_name not in desc_positions:
                continue
            bstart, bpos = desc_positions[set_name]
            block = new_content[bstart:bpos]
            desc_match = re.search(r'DESCRIPTION="([^"]*)"', block)
            if desc_match:
                old_desc = desc_match.group(1)
                new_block = block[:desc_match.start()] + f'DESCRIPTION="{new_desc}"' + block[desc_match.end():]
            else:
                old_desc = None
                brace_pos = block.find('{')
                if brace_pos >= 0:
                    insert_pos = brace_pos + 1
                    while insert_pos < len(block) and block[insert_pos] in ' \t\r\n':
                        insert_pos += 1
                    new_block = block[:insert_pos] + f'\r\n  DESCRIPTION="{new_desc}"' + block[insert_pos:]
                else:
                    continue
            modified_desc_blocks[set_name] = (bstart, bpos, new_block)
            desc_count += 1
            if old_desc is not None:
                log(f"  {set_name}: DESCRIPTION \"{old_desc}\" -> \"{new_desc}\"")
            else:
                log(f"  {set_name}: DESCRIPTION added \"{new_desc}\"")

        sorted_desc = sorted(modified_desc_blocks.items(), key=lambda x: x[1][0], reverse=True)
        for idx, (set_name, (bstart, bpos, new_block)) in enumerate(sorted_desc):
            if idx % 10 == 0:
                progress(30 + int(8 * idx / max(1, len(sorted_desc))),
                         f'Reassembling descriptions... ({idx}/{len(sorted_desc)})')
            log(f"  [{idx+1}/{len(sorted_desc)}] DESCRIPTION {set_name}")
            new_content = new_content[:bstart] + new_block + new_content[bpos:]
        progress(38, f'Updated {len(sorted_desc)} descriptions')

    # Add new namesets
    progress(40, 'Adding new namesets...')
    added_count = 0
    ns_total = len(new_namesets)
    for idx, ns in enumerate(new_namesets):
        if ns_total > 0 and idx % max(1, ns_total // 20) == 0:
            pct = 40 + int(10 * idx / ns_total)
            progress(pct, f'Adding new namesets... ({idx}/{ns_total})')
        name = ns['name']
        category = ns['category']
        entries = ns['entries']
        description = ns.get('description', '')

        block_lines = [f'ENUMERATION_SET NAME="{name}" FIXED=F']
        block_lines.append(f' user="Emerson" time=0')
        block_lines.append('{')
        if description:
            block_lines.append(f'  DESCRIPTION="{description}"')
        if category:
            block_lines.append(f'  CATEGORY="{category}"')
        for entry in entries:
            block_lines.append(f'  ENTRY VALUE={entry["value"]} NAME="{entry["name"]}" {{ }}')
        if entries:
            block_lines.append(f'  DEFAULT_VALUE={entries[0]["value"]}')
        block_lines.append('}')
        block = '\r\n'.join(block_lines)

        insert_idx = new_content.find('FUNCTION_BLOCK_DEFINITIONS')
        if insert_idx < 0:
            insert_idx = new_content.find('FUNCTION_BLOCK ')
        if insert_idx < 0:
            locale_end = new_content.find('\n}\n', new_content.find('LOCALE'))
            if locale_end > 0:
                insert_idx = locale_end + 3
            else:
                insert_idx = len(new_content)
        new_content = new_content[:insert_idx] + block + '\r\n' + new_content[insert_idx:]
        added_count += 1
        log(f"  Added new nameset: {name}")

    # Replace STRING_VALUE references (batch: single regex pass)
    progress(50, 'Replacing STRING_VALUE references...')
    sv_count = 0
    sv_replacements = {}
    for set_name, old_value, new_value in sv_changes:
        new_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_value)
        new_val_name = new_name_m.group(1) if new_name_m else new_value
        old_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_value)
        old_val_name = old_name_m.group(1) if old_name_m else old_value
        if old_val_name != new_val_name:
            sv_replacements[(set_name, old_val_name)] = new_val_name

    if sv_replacements:
        sv_by_set = {}
        for (set_name, old_val), new_val in sv_replacements.items():
            sv_by_set.setdefault(set_name, {})[old_val] = new_val

        log(f"  {len(sv_by_set)} sets, {len(sv_replacements)} values to replace")
        progress(50, f'Replacing {len(sv_replacements)} STRING_VALUE refs...')

        # Pass 1: build set-name-keyed map for fast lookup in repl callback
        sv_pair_map = {}
        sv_set_parts = []
        for set_name, val_map in sv_by_set.items():
            escaped_set = re.escape(set_name)
            old_vals_escaped = sorted([re.escape(v) for v in val_map.keys()], key=len, reverse=True)
            vals_alt = '|'.join(old_vals_escaped)
            sv_set_parts.append(f'SET="{escaped_set}"[^}}]*?STRING_VALUE="({vals_alt})"')
            for k, v in val_map.items():
                sv_pair_map[(set_name, k)] = v

        # Pass 2: build reverse map val_name -> set_name (for non-colliding values)
        # and a collision set for values that appear in multiple sets
        sv_val_to_sets = {}
        for (sn, val) in sv_pair_map:
            sv_val_to_sets.setdefault(val, set()).add(sn)
        sv_colliding = {v for v, sn_set in sv_val_to_sets.items() if len(sn_set) > 1}

        sv_combined = '|'.join(sv_set_parts)
        def sv_repl(m, vm=sv_pair_map, coll=sv_colliding):
            val = next(g for g in m.groups() if g is not None)
            full = m.group(0)
            if val not in coll:
                new_val = vm[next(k for k in vm if k[1] == val)]
            else:
                set_m = re.search(r'SET="([^"]*)"', full)
                sn = set_m.group(1)
                new_val = vm[(sn, val)]
            # Replace only inside STRING_VALUE="...", not in SET name
            return re.sub(r'(STRING_VALUE=")' + re.escape(val) + r'(")',
                          f'\\g<1>{new_val}\\2', full, count=1)
        new_content, sv_count = re.subn(sv_combined, sv_repl, new_content, flags=re.DOTALL)
        log(f"  STRING_VALUE: {sv_count} replacements")
        progress(70, f'Replaced {sv_count} STRING_VALUE refs')

    # Replace expression references (batch: single regex pass)
    progress(70, 'Replacing expression references...')
    expr_count = 0
    expr_replacements = {}
    for set_name, old_value, new_value in expr_changes:
        new_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_value)
        if new_name_m:
            new_val_name = new_name_m.group(1)
        else:
            nm = re.match(r'NAME="([^"]*)"', new_value)
            new_val_name = nm.group(1) if nm else new_value

        old_name_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_value)
        old_val_name = old_name_m.group(1) if old_name_m else old_value

        actual_old_name = old_val_name
        if set_name in lib_enum_defs:
            for entry in lib_enum_defs[set_name]['entries']:
                em = re.match(r'VALUE=(\d+)\s+NAME="([^"]*)"', entry)
                if em and em.group(2) == old_val_name:
                    break
                new_vm = re.match(r'VALUE=(\d+)', new_value)
                if new_vm and em and em.group(1) == new_vm.group(1):
                    actual_old_name = em.group(2)
                    break

        if actual_old_name != new_val_name:
            expr_replacements[(set_name, actual_old_name)] = new_val_name

    if expr_replacements:
        expr_by_set = {}
        for (set_name, old_val), new_val in expr_replacements.items():
            expr_by_set.setdefault(set_name, {})[old_val] = new_val

        log(f"  {len(expr_by_set)} sets, {len(expr_replacements)} expression values to replace")
        progress(70, f'Replacing {len(expr_replacements)} expression refs...')

        expr_all_val_map = {}
        expr_set_parts = []
        for set_name, val_map in expr_by_set.items():
            escaped_set = re.escape(set_name)
            old_vals_escaped = sorted([re.escape(v) for v in val_map.keys()], key=len, reverse=True)
            vals_alt = '|'.join(old_vals_escaped)
            expr_set_parts.append(f"'({escaped_set}):({vals_alt})'")
            for k, v in val_map.items():
                expr_all_val_map[(set_name, k)] = v

        expr_combined = '|'.join(expr_set_parts)
        def expr_repl(m, vm=expr_all_val_map):
            # Each alternative contributes 2 groups; find the non-None pair
            groups = m.groups()
            set_name = None
            val_name = None
            for i in range(0, len(groups), 2):
                if groups[i] is not None:
                    set_name = groups[i]
                    val_name = groups[i + 1]
                    break
            return f"'{set_name}:{vm[(set_name, val_name)]}'"
        new_content, expr_count = re.subn(expr_combined, expr_repl, new_content)
        log(f"  Expression refs: {expr_count} replacements")
        progress(85, f'Replaced {expr_count} expression refs')

    # Additional pass: replace expression values from nameset_changes (batch)
    progress(85, 'Replacing expression refs from nameset changes...')
    if nameset_changes:
        ns_expr_map = {}
        for (ns_name, old_entry), new_entry in nameset_changes.items():
            old_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', old_entry)
            new_m = re.match(r'VALUE=\d+\s+NAME="([^"]*)"', new_entry)
            if old_m and new_m and old_m.group(1) != new_m.group(1):
                ns_expr_map[(ns_name, old_m.group(1))] = new_m.group(1)

        log(f"  Expression ref candidates from nameset changes: {len(ns_expr_map)}")
        if ns_expr_map:
            ns_by_set = {}
            for (ns_name, old_val), new_val in ns_expr_map.items():
                ns_by_set.setdefault(ns_name, {})[old_val] = new_val

            ns_all_val_map = {}
            ns_set_parts = []
            for ns_name, val_map in ns_by_set.items():
                escaped_set = re.escape(ns_name)
                old_vals_escaped = sorted([re.escape(v) for v in val_map.keys()], key=len, reverse=True)
                vals_alt = '|'.join(old_vals_escaped)
                ns_set_parts.append(f"'({escaped_set}):({vals_alt})'")
                for k, v in val_map.items():
                    ns_all_val_map[(ns_name, k)] = v

            ns_combined = '|'.join(ns_set_parts)
            def ns_repl(m, vm=ns_all_val_map):
                groups = m.groups()
                set_name = None
                val_name = None
                for i in range(0, len(groups), 2):
                    if groups[i] is not None:
                        set_name = groups[i]
                        val_name = groups[i + 1]
                        break
                return f"'{set_name}:{vm[(set_name, val_name)]}'"
            new_content, ns_count = re.subn(ns_combined, ns_repl, new_content)
            expr_count += ns_count
            log(f"  Nameset expression refs: {ns_count} replacements")
        else:
            log("  No expression-style name changes found in nameset changes (all use VALUE=N NAME= format)")
    else:
        log("  No nameset changes to process")

    # Replace alarm values
    alarm_count = 0
    if alarm_changes:
        new_content, alarm_count = replace_alarm_values(new_content, alarm_changes)
        log(f"  Replaced {alarm_count} alarm fields")

    # Replace PRIORITY_NAME values
    priority_count = 0
    if priority_changes:
        new_content, priority_count = replace_priority_names(new_content, priority_changes)
        log(f"  Replaced {priority_count} PRIORITY_NAME values")

    # Write output
    progress(90, 'Writing output...')
    log(f"\nWriting output: {output_path}")
    write_fhx(output_path, new_content)

    progress(100, 'Done')
    log(f"  Done! Replaced {replace_count} ENTRY NAME, Updated {desc_count} descriptions, "
        f"Added {added_count} new namesets, Replaced {sv_count} STRING_VALUE, "
        f"Replaced {expr_count} expression refs, Replaced {alarm_count} alarm fields, "
        f"Translated {priority_count} PRIORITY_NAME")

    return replace_count + desc_count + added_count + sv_count + expr_count + alarm_count + priority_count

# ============================================================
# Subprocess entry point (JSON communication, no GUI)
# ============================================================
def _subprocess_generate(lib_path, setup_path, nameset_changes, new_namesets, desc_changes,
                         sv_changes, expr_changes, output_path, alarm_changes, priority_changes):
    """Entry point for subprocess: run generate_new_lib_fhx in a separate process."""
    # nameset_changes arrives as list of [[set_name, old_entry], new_entry] from JSON
    if isinstance(nameset_changes, list):
        nameset_changes = {tuple(k): v for k, v in nameset_changes}
    def log_cb(msg):
        print(json.dumps({'type': 'log', 'msg': msg}), flush=True)
    def prog_cb(pct, text):
        print(json.dumps({'type': 'progress', 'pct': pct, 'text': text}), flush=True)
    try:
        count = generate_new_lib_fhx(
            lib_path, setup_path, nameset_changes, new_namesets, desc_changes,
            sv_changes, expr_changes, output_path,
            alarm_changes=alarm_changes, priority_changes=priority_changes,
            log_callback=log_cb, progress_callback=prog_cb)
        print(json.dumps({'type': 'done', 'ok': True, 'count': count}), flush=True)
    except Exception as e:
        print(json.dumps({'type': 'done', 'ok': False, 'error': str(e)}), flush=True)

# ============================================================
# CLI helpers
# ============================================================
_cli_start_time = None

def _supports_unicode():
    """Check if the terminal supports Unicode block characters."""
    enc = getattr(sys.stdout, 'encoding', '') or ''
    return enc.lower() not in ('gbk', 'cp936', 'gb2312', 'gb18030', 'ascii', '')

def cli_progress(pct, text=''):
    """CLI progress bar with ETA."""
    import sys as _sys
    global _cli_start_time
    if _cli_start_time is None:
        _cli_start_time = time.time()
    bar_len = 40
    filled = int(bar_len * pct / 100)
    if not hasattr(cli_progress, '_unicode'):
        cli_progress._unicode = _supports_unicode()
    if cli_progress._unicode:
        bar = '█' * filled + '░' * (bar_len - filled)
    else:
        bar = '#' * filled + '-' * (bar_len - filled)
    elapsed = time.time() - _cli_start_time
    if pct > 0 and pct < 100:
        eta = elapsed * (100 - pct) / pct
        time_str = f'  {elapsed:.1f}s  ETA:{eta:.0f}s'
    elif pct >= 100:
        time_str = f'  Done in {elapsed:.1f}s'
        _cli_start_time = None
        _sys.stdout.write(f'\r  [{bar}] {pct:3d}% {text}{time_str}' + ' ' * 10 + '\n')
        _sys.stdout.flush()
        return
    else:
        time_str = f'  {elapsed:.1f}s'
    _sys.stdout.write(f'\r  [{bar}] {pct:3d}% {text}{time_str}')
    _sys.stdout.flush()
